# -*- coding: utf-8 -*-
"""اختبارات طبقة التصدير الموحدة ui/exporters.py"""
import os
import tempfile
import unittest
from unittest.mock import patch

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook

from ui import exporters


class TestExporters(unittest.TestCase):

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def _path(self, name):
        return os.path.join(self.tmpdir, name)

    def test_style_header_row(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["A", "B", "C"])
        exporters.style_header_row(ws)
        ws.append([1, 2, 3])
        wb.save(self._path("h.xlsx"))
        wb2 = load_workbook(self._path("h.xlsx"))
        ws2 = wb2.active
        self.assertTrue(ws2["A1"].font.bold)
        self.assertEqual(ws2["A1"].fill.start_color.rgb, "002980B9")
        self.assertFalse(ws2["A2"].font.bold)

    def test_style_header_row_custom_row(self):
        wb = Workbook()
        ws = wb.active
        ws.append(["skip"])
        ws.append(["H1", "H2"])
        exporters.style_header_row(ws, row=2)
        self.assertTrue(ws["A2"].font.bold)
        self.assertFalse(ws["A1"].font.bold)

    def test_write_charts_pdf(self):
        figs = []
        for _ in range(2):
            fig, ax = plt.subplots()
            ax.plot([1, 2, 3])
            plt.close(fig)
            figs.append(fig)
        path = self._path("charts.pdf")
        exporters.write_charts_pdf(path, figs)
        from pypdf import PdfReader
        self.assertEqual(len(PdfReader(path).pages), 2)

    def test_ask_save_path_chosen(self):
        from PyQt5.QtWidgets import QApplication
        app = QApplication.instance() or QApplication([])
        with patch("ui.exporters.QFileDialog.getSaveFileName",
                   return_value=("C:/out.pdf", "PDF (*.pdf)")):
            self.assertEqual(exporters.ask_save_path(None, "cap", "d.pdf", "PDF (*.pdf)"),
                             "C:/out.pdf")

    def test_ask_save_path_cancelled(self):
        with patch("ui.exporters.QFileDialog.getSaveFileName",
                   return_value=("", "")):
            self.assertIsNone(exporters.ask_save_path(None, "cap", "d.pdf", "PDF (*.pdf)"))

    def test_new_workbook_and_add_sheet(self):
        wb = exporters.new_workbook()
        ws = exporters.add_excel_sheet(
            wb, "Profitability",
            ["Name", "Revenue", "Margin"],
            [["A", 1000, 0.2], ["B", 2000, 0.3]],
        )
        self.assertEqual(ws.title, "Profitability")
        self.assertEqual(ws.max_row, 3)          # رأس + سطران
        self.assertEqual(ws.max_column, 3)
        self.assertTrue(ws["A1"].font.bold)      # رأس ملوّن
        self.assertEqual(ws["A2"].value, "A")
        self.assertEqual(ws["C3"].value, 0.3)

    def test_add_sheet_multiple_uses_active_first(self):
        wb = exporters.new_workbook()
        ws1 = exporters.add_excel_sheet(wb, "First", ["H"], [["1"]])
        ws2 = exporters.add_excel_sheet(wb, "Second", ["H"], [["2"]])
        ws3 = exporters.add_excel_sheet(wb, "Third", ["H"], [["3"]])
        self.assertEqual(wb.sheetnames, ["First", "Second", "Third"])
        self.assertEqual(ws1.value(1, 1) if hasattr(ws1, "value") else ws1["A2"].value, "1")
        self.assertEqual(ws2["A2"].value, "2")
        self.assertEqual(ws3["A2"].value, "3")
        # كل رأس ملوّن بنفس اللون الافتراضي
        self.assertEqual(ws2["A1"].fill.start_color.rgb, "001F4E79")

    def test_add_sheet_roundtrip_save_load(self):
        wb = exporters.new_workbook()
        exporters.add_excel_sheet(wb, "Data", ["A", "B"], [[1, 9], [2, 8]])
        path = self._path("roundtrip.xlsx")
        wb.save(path)
        wb2 = load_workbook(path)
        ws2 = wb2["Data"]
        self.assertEqual(ws2["A2"].value, 1)
        self.assertEqual(ws2["B3"].value, 8)
        self.assertTrue(ws2["A1"].font.bold)

    def test_add_sheet_custom_header_fill(self):
        wb = exporters.new_workbook()
        ws = exporters.add_excel_sheet(
            wb, "Custom", ["H"], [[]], header_fill="ABCDEF")
        self.assertEqual(ws["A1"].fill.start_color.rgb, "00ABCDEF")


if __name__ == "__main__":
    unittest.main()
