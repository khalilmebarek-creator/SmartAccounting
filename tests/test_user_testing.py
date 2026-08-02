# اختبارات محرك اختبار المستخدمين الحقيقيين
# ===========================================
# مجموعات المستخدمين (4) × سيناريوهات (5) × تصنيفات الملاحظات (5)
# + درجة رضا المستخدم + تقارير + JSON/DB/Excel/PDF/CSV

import json
import os
import sqlite3
import sys
import tempfile
import unittest
from unittest import mock

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import config
from database import db_connection
from modules.user_testing import (
    UserTestingEngine, USER_GROUPS, SCENARIOS, FEEDBACK_CATEGORIES,
    PRIORITIES, STATUSES, satisfaction_level,
)

try:
    import openpyxl  # noqa: F401
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import fpdf  # noqa: F401
    HAS_FPDF = True
except ImportError:
    HAS_FPDF = False


class UserTestingEngineTest(unittest.TestCase):
    def setUp(self):
        self.engine = UserTestingEngine()
        self.session = self.engine.create_session(
            name="جلسة تجريبية", tester_name="أحمد",
            user_group="accountant", scenario="data_entry",
            environment="Windows 11",
        )
        self.sid = self.session["id"]

    # ---------- الجلسات ----------

    def test_create_session_sets_fields(self):
        session = self.engine.create_session(
            "جلسة 2", tester_name="ليلى", user_group="manager",
            scenario="report_generation", environment="macOS",
        )
        self.assertEqual(session["name"], "جلسة 2")
        self.assertEqual(session["tester_name"], "ليلى")
        self.assertEqual(session["user_group"], "manager")
        self.assertEqual(session["scenario"], "report_generation")
        self.assertEqual(session["environment"], "macOS")
        self.assertEqual(session["feedback"], [])
        self.assertNotEqual(session["id"], self.sid)
        self.assertTrue(session["created_at"])

    def test_create_session_requires_name(self):
        with self.assertRaises(ValueError):
            self.engine.create_session("   ")

    def test_create_session_rejects_invalid_group(self):
        with self.assertRaises(ValueError):
            self.engine.create_session("x", user_group="auditor")

    def test_create_session_rejects_invalid_scenario(self):
        with self.assertRaises(ValueError):
            self.engine.create_session("x", scenario="typing")

    def test_get_session_missing_returns_none(self):
        self.assertIsNone(self.engine.get_session("999"))

    def test_delete_session(self):
        self.assertTrue(self.engine.delete_session(self.sid))
        self.assertFalse(self.engine.delete_session(self.sid))
        self.assertIsNone(self.engine.get_session(self.sid))

    def test_list_sessions(self):
        self.engine.create_session("ثانية", user_group="cfo")
        self.assertEqual(len(self.engine.list_sessions()), 2)

    # ---------- الملاحظات والتحقق ----------

    def test_add_feedback_defaults_to_session_group_scenario(self):
        feedback = self.engine.add_feedback(
            self.sid, category="usability", comment="واضحة", rating=4,
        )
        self.assertEqual(feedback["user_group"], "accountant")
        self.assertEqual(feedback["scenario"], "data_entry")
        self.assertEqual(feedback["priority"], "medium")
        self.assertEqual(feedback["status"], "open")
        self.assertEqual(feedback["rating"], 4)
        self.assertTrue(feedback["id"])

    def test_add_feedback_validation(self):
        cases = [
            dict(category="weird", comment="x", rating=3),
            dict(category="bugs", comment="x", rating=0),
            dict(category="bugs", comment="x", rating=6),
            dict(category="bugs", comment="x", rating="abc"),
            dict(category="bugs", comment="  ", rating=3),
            dict(category="bugs", comment="x", rating=3, priority="urgent"),
            dict(category="bugs", comment="x", rating=3, status="done"),
            dict(category="bugs", comment="x", rating=3, user_group="admin"),
            dict(category="bugs", comment="x", rating=3, scenario="typing"),
        ]
        for kwargs in cases:
            with self.assertRaises(ValueError):
                self.engine.add_feedback(self.sid, **kwargs)

    def test_add_feedback_missing_session(self):
        with self.assertRaises(KeyError):
            self.engine.add_feedback("999", "bugs", "x", 3)

    def test_update_feedback(self):
        feedback = self.engine.add_feedback(
            self.sid, "bugs", "عطل", 2, priority="high",
        )
        self.engine.update_feedback(
            self.sid, feedback["id"], status="resolved", priority="critical",
        )
        updated = self.engine.get_session(self.sid)["feedback"][0]
        self.assertEqual(updated["status"], "resolved")
        self.assertEqual(updated["priority"], "critical")

    def test_update_feedback_invalid(self):
        feedback = self.engine.add_feedback(self.sid, "bugs", "عطل", 3)
        with self.assertRaises(ValueError):
            self.engine.update_feedback(self.sid, feedback["id"], rating=7)
        with self.assertRaises(KeyError):
            self.engine.update_feedback(self.sid, feedback["id"], bogus=1)

    def test_update_feedback_missing(self):
        with self.assertRaises(KeyError):
            self.engine.update_feedback(self.sid, "nope", status="resolved")

    def test_delete_feedback(self):
        feedback = self.engine.add_feedback(self.sid, "bugs", "عطل", 3)
        self.assertTrue(self.engine.delete_feedback(self.sid, feedback["id"]))
        self.assertFalse(self.engine.delete_feedback(self.sid, feedback["id"]))
        self.assertEqual(self.engine.get_session(self.sid)["feedback"], [])

    def test_list_feedback_filters(self):
        self.engine.add_feedback(self.sid, "bugs", "عطل", 2, user_group="manager")
        self.engine.add_feedback(self.sid, "features", "طلب", 5)
        bugs = self.engine.list_feedback(self.sid, category="bugs")
        self.assertEqual(len(bugs), 1)
        self.assertEqual(bugs[0]["comment"], "عطل")
        managers = self.engine.list_feedback(self.sid, user_group="manager")
        self.assertEqual(len(managers), 1)
        open_items = self.engine.list_feedback(self.sid, status="open")
        self.assertEqual(len(open_items), 2)

    def test_list_feedback_missing_session(self):
        self.assertEqual(self.engine.list_feedback("999"), [])

    # ---------- درجة الرضا ----------

    def test_satisfaction_score_empty(self):
        score = self.engine.satisfaction_score(self.sid)
        self.assertEqual(score["overall"], 0.0)
        self.assertEqual(score["count"], 0)
        self.assertEqual(score["level"], "poor")

    def test_satisfaction_score_missing_session(self):
        score = self.engine.satisfaction_score("999")
        self.assertEqual(score["overall"], 0.0)

    def test_satisfaction_score_mean_and_breakdown(self):
        self.engine.add_feedback(self.sid, "usability", "أ", 5)
        self.engine.add_feedback(self.sid, "bugs", "ب", 3)
        self.engine.add_feedback(self.sid, "usability", "ج", 4, user_group="manager")
        score = self.engine.satisfaction_score(self.sid)
        self.assertEqual(score["overall"], 4.0)
        self.assertEqual(score["count"], 3)
        self.assertEqual(score["by_category"], {"bugs": 3.0, "usability": 4.5})
        self.assertEqual(score["by_user_group"], {"accountant": 4.0, "manager": 4.0})
        self.assertEqual(score["by_scenario"], {"data_entry": 4.0})

    def test_satisfaction_levels_thresholds(self):
        self.assertEqual(satisfaction_level(5.0), "excellent")
        self.assertEqual(satisfaction_level(4.5), "excellent")
        self.assertEqual(satisfaction_level(3.9), "good")
        self.assertEqual(satisfaction_level(3.0), "average")
        self.assertEqual(satisfaction_level(2.0), "poor")
        self.assertEqual(satisfaction_level(0.0), "poor")

    # ---------- التقارير ----------

    def test_feedback_report_structure(self):
        self.engine.add_feedback(self.sid, "bugs", "عطل", 2, priority="critical")
        self.engine.add_feedback(self.sid, "features", "طلب ميزة", 5)
        self.engine.add_feedback(self.sid, "suggestions", "اقتراح", 4)
        report = self.engine.feedback_report(self.sid)
        self.assertEqual(report["session"]["name"], "جلسة تجريبية")
        self.assertEqual(report["total_feedback"], 3)
        self.assertEqual(report["counts_by_category"]["bugs"], 1)
        self.assertEqual(report["counts_by_status"]["open"], 3)
        self.assertEqual(report["counts_by_priority"]["critical"], 1)
        self.assertEqual(report["open_issues"], 1)
        self.assertEqual(report["enhancement_requests"], 2)
        self.assertEqual(len(report["feedback"]), 3)
        self.assertEqual(report["satisfaction"]["overall"], round(11 / 3, 2))

    def test_feedback_report_empty_session(self):
        report = self.engine.feedback_report(self.sid)
        self.assertEqual(report["total_feedback"], 0)
        self.assertEqual(report["open_issues"], 0)

    def test_feedback_report_missing_session(self):
        self.assertEqual(self.engine.feedback_report("999"), {})

    def test_issue_list_only_bugs_sorted_by_priority(self):
        self.engine.add_feedback(self.sid, "bugs", "عطل خفيف", 3, priority="low")
        self.engine.add_feedback(self.sid, "features", "ميزة", 5, priority="high")
        self.engine.add_feedback(self.sid, "bugs", "عطل خطير", 2, priority="critical")
        self.engine.add_feedback(self.sid, "bugs", "عطل متوسط", 3, priority="medium")
        self.engine.add_feedback(self.sid, "suggestions", "اقتراح", 4)
        issues = self.engine.issue_list(self.sid)
        self.assertEqual([i["comment"] for i in issues],
                         ["عطل خطير", "عطل متوسط", "عطل خفيف"])

    def test_issue_list_missing_session(self):
        self.assertEqual(self.engine.issue_list("999"), [])

    def test_enhancement_requests_filters(self):
        self.engine.add_feedback(self.sid, "bugs", "عطل", 2, priority="critical")
        self.engine.add_feedback(self.sid, "features", "ميزة أ", 5, priority="medium")
        self.engine.add_feedback(self.sid, "suggestions", "اقتراح ب", 4, priority="high")
        self.engine.add_feedback(self.sid, "usability", "سهولة", 4)
        requests = self.engine.enhancement_requests(self.sid)
        self.assertEqual([i["comment"] for i in requests], ["اقتراح ب", "ميزة أ"])

    def test_summary_text_contains_score(self):
        self.engine.add_feedback(self.sid, "bugs", "عطل", 3)
        text = self.engine.summary_text(self.sid)
        self.assertIn("درجة الرضا: 3.0/5", text)
        self.assertIn("جلسة تجريبية", text)
        self.assertIn("عطل", text)

    def test_summary_text_missing_session(self):
        self.assertTrue(self.engine.summary_text("999"))

    # ---------- JSON ----------

    def test_json_roundtrip(self):
        self.engine.add_feedback(self.sid, "bugs", "عطل", 2, priority="critical")
        self.engine.add_feedback(self.sid, "usability", "جيد", 4)
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "sessions.json")
            self.assertTrue(self.engine.export_json(path))
            fresh = UserTestingEngine()
            self.assertEqual(fresh.import_json(path), 1)
            loaded = fresh.get_session(self.sid)
            self.assertEqual(loaded["name"], "جلسة تجريبية")
            self.assertEqual(len(loaded["feedback"]), 2)
            self.assertEqual(
                fresh.satisfaction_score(self.sid)["overall"], 3.0
            )

    def test_import_json_missing_file(self):
        with tempfile.TemporaryDirectory() as tmp:
            self.assertEqual(
                self.engine.import_json(os.path.join(tmp, "none.json")), 0
            )

    def test_import_json_bad_payload(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "bad.json")
            with open(path, "w", encoding="utf-8") as f:
                json.dump({"foo": 1}, f)
            self.assertEqual(self.engine.import_json(path), 0)
            with open(path, "w", encoding="utf-8") as f:
                f.write("{not json")
            self.assertEqual(self.engine.import_json(path), 0)

    def test_import_session_generates_new_id_on_conflict(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "s.json")
            self.engine.export_json(path)
            engine2 = UserTestingEngine()
            engine2.import_json(path)
            engine3 = UserTestingEngine()
            engine3.import_json(path)
            engine3.import_json(path)
            self.assertEqual(len(engine3.list_sessions()), 2)

    # ---------- CSV ----------

    def test_export_csv(self):
        self.engine.add_feedback(self.sid, "bugs", "عطل", 2, title="تعطل الإدخال")
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "feedback.csv")
            self.assertTrue(self.engine.export_csv(path))
            with open(path, "r", encoding="utf-8-sig") as f:
                content = f.read()
            self.assertIn("تعطل الإدخال", content)
            self.assertIn("bugs", content)

    def test_export_csv_no_sessions(self):
        engine = UserTestingEngine()
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "feedback.csv")
            self.assertTrue(engine.export_csv(path))

    # ---------- Excel ----------

    @unittest.skipUnless(HAS_OPENPYXL, "openpyxl not installed")
    def test_export_excel(self):
        self.engine.add_feedback(self.sid, "bugs", "عطل", 2)
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "feedback.xlsx")
            self.assertTrue(self.engine.export_excel(path))
            self.assertTrue(os.path.exists(path))
            self.assertGreater(os.path.getsize(path), 0)
            from openpyxl import load_workbook
            wb = load_workbook(path)
            self.assertIn("Feedback", wb.sheetnames)
            self.assertIn("Summary", wb.sheetnames)

    def test_export_excel_no_openpyxl(self):
        with mock.patch.dict("sys.modules", {"openpyxl": None}):
            with tempfile.TemporaryDirectory() as tmp:
                path = os.path.join(tmp, "feedback.xlsx")
                self.assertFalse(self.engine.export_excel(path))

    # ---------- PDF ----------

    @unittest.skipUnless(HAS_FPDF, "fpdf not installed")
    def test_export_pdf(self):
        self.engine.add_feedback(self.sid, "suggestions", "اقتراح", 5)
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "feedback.pdf")
            self.assertTrue(self.engine.export_pdf(path))
            self.assertTrue(os.path.exists(path))
            self.assertGreater(os.path.getsize(path), 0)

    def test_export_pdf_no_fpdf(self):
        with mock.patch.dict("sys.modules", {"fpdf": None}):
            with tempfile.TemporaryDirectory() as tmp:
                path = os.path.join(tmp, "feedback.pdf")
                self.assertFalse(self.engine.export_pdf(path))

    # ---------- قاعدة البيانات ----------

    def test_db_roundtrip(self):
        self.engine.add_feedback(self.sid, "bugs", "عطل", 2, priority="high")
        original = config.DATABASE_PATH
        tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        tmp_db.close()
        try:
            config.DATABASE_PATH = tmp_db.name
            db_connection.close_pool()
            self.assertTrue(self.engine.save_session_db(self.sid))
            ids = self.engine.list_session_ids_db()
            self.assertEqual(len(ids), 1)
            self.assertEqual(ids[0]["id"], self.sid)

            fresh = UserTestingEngine()
            self.assertTrue(fresh.load_session_db(self.sid))
            loaded = fresh.get_session(self.sid)
            self.assertEqual(loaded["name"], "جلسة تجريبية")
            self.assertEqual(len(loaded["feedback"]), 1)

            self.assertTrue(fresh.delete_session_db(self.sid))
            self.assertEqual(fresh.list_session_ids_db(), [])
        finally:
            config.DATABASE_PATH = original
            db_connection.close_pool()
            try:
                os.remove(tmp_db.name)
                os.remove(tmp_db.name + "-wal")
                os.remove(tmp_db.name + "-shm")
            except OSError:
                pass

    def test_save_session_db_missing_session(self):
        self.assertFalse(self.engine.save_session_db("999"))

    def test_load_session_db_when_table_missing(self):
        original = config.DATABASE_PATH
        tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        tmp_db.close()
        try:
            config.DATABASE_PATH = tmp_db.name
            db_connection.close_pool()
            self.assertFalse(self.engine.load_session_db(self.sid))
        finally:
            config.DATABASE_PATH = original
            db_connection.close_pool()
            try:
                os.remove(tmp_db.name)
                os.remove(tmp_db.name + "-wal")
                os.remove(tmp_db.name + "-shm")
            except OSError:
                pass

    # ---------- بيانات تجريبية ----------

    def test_build_demo_data(self):
        engine = UserTestingEngine()
        count = engine.build_demo_data()
        self.assertEqual(count, 4)
        self.assertEqual(len(engine.list_sessions()), 4)
        groups = {s["user_group"] for s in engine.list_sessions()}
        self.assertEqual(groups, set(USER_GROUPS))
        for session in engine.list_sessions():
            self.assertTrue(session["feedback"])
        engine.build_demo_data()
        self.assertEqual(len(engine.list_sessions()), 4)

    def test_constants(self):
        self.assertEqual(len(USER_GROUPS), 4)
        self.assertEqual(len(SCENARIOS), 5)
        self.assertEqual(len(FEEDBACK_CATEGORIES), 5)
        self.assertEqual(len(PRIORITIES), 4)
        self.assertEqual(len(STATUSES), 4)


class UserTestingCoverageTest(unittest.TestCase):
    """استكمال الفروع الحافة للوصول إلى 100% تغطية."""

    def setUp(self):
        self.engine = UserTestingEngine()
        self.session = self.engine.create_session(
            "تغطية", tester_name="اختبار", user_group="manager",
            scenario="analysis",
        )
        self.sid = self.session["id"]

    # ---------- _atomic_write ----------

    def test_atomic_write_replace_failure_rethrows(self):
        with mock.patch("modules.user_testing.os.replace",
                        side_effect=OSError("boom")):
            with tempfile.TemporaryDirectory() as tmp:
                with self.assertRaises(OSError):
                    self.engine.export_json(os.path.join(tmp, "x.json"))

    def test_atomic_write_remove_failure_swallowed(self):
        with mock.patch("modules.user_testing.os.replace",
                        side_effect=OSError("boom")), \
             mock.patch("modules.user_testing.os.remove",
                        side_effect=OSError("remove fail")):
            with tempfile.TemporaryDirectory() as tmp:
                with self.assertRaises(OSError):
                    self.engine.export_json(os.path.join(tmp, "x.json"))

    # ---------- _wrap_line ----------

    def test_wrap_line_empty(self):
        from modules.user_testing import _wrap_line
        self.assertEqual(_wrap_line(_FakePdf(10), "", 100), [""])

    def test_wrap_line_fits(self):
        from modules.user_testing import _wrap_line
        self.assertEqual(_wrap_line(_FakePdf(10), "short", 100), ["short"])

    def test_wrap_line_wraps_long(self):
        from modules.user_testing import _wrap_line
        pdf = _FakePdf(10)
        self.assertEqual(_wrap_line(pdf, "one two three four", 25),
                         ["one two", "three four"])

    # ---------- update_feedback validation branches ----------

    def test_update_feedback_invalid_category(self):
        feedback = self.engine.add_feedback(self.sid, "bugs", "x", 3)
        with self.assertRaises(ValueError):
            self.engine.update_feedback(self.sid, feedback["id"], category="nope")

    def test_update_feedback_invalid_user_group(self):
        feedback = self.engine.add_feedback(self.sid, "bugs", "x", 3)
        with self.assertRaises(ValueError):
            self.engine.update_feedback(self.sid, feedback["id"], user_group="nope")

    def test_update_feedback_invalid_scenario(self):
        feedback = self.engine.add_feedback(self.sid, "bugs", "x", 3)
        with self.assertRaises(ValueError):
            self.engine.update_feedback(self.sid, feedback["id"], scenario="nope")

    def test_delete_feedback_missing_session(self):
        with self.assertRaises(KeyError):
            self.engine.delete_feedback("999", "x")

    def test_update_feedback_missing_session(self):
        with self.assertRaises(KeyError):
            self.engine.update_feedback("999", "x", status="resolved")

    def test_enhancement_requests_missing_session(self):
        self.assertEqual(self.engine.enhancement_requests("999"), [])

    # ---------- _import_session branches ----------

    def test_import_session_without_name(self):
        self.assertFalse(self.engine._import_session({"id": "9", "foo": 1}))
        self.assertFalse(self.engine._import_session("not a dict"))

    def test_import_session_invalid_rating(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "bad.json")
            with open(path, "w", encoding="utf-8") as f:
                json.dump({"sessions": [{"name": "x", "feedback": [{"rating": 9}]}]}, f)
            self.assertEqual(self.engine.import_json(path), 0)

    def test_import_session_non_numeric_id(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "str.json")
            with open(path, "w", encoding="utf-8") as f:
                json.dump({"sessions": [{"id": "abc", "name": "x"}]}, f)
            self.assertEqual(self.engine.import_json(path), 1)
            self.assertEqual(len(self.engine.list_sessions()), 2)
            self.assertIsNotNone(self.engine.get_session("abc"))

    # ---------- DB branches ----------

    def test_db_connection_error_paths(self):
        original = config.DATABASE_PATH
        tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        tmp_db.close()
        try:
            config.DATABASE_PATH = tmp_db.name
            db_connection.close_pool()
            # جدول غير موجود → قوائم/حذف يعيدان الافتراضي
            self.assertEqual(self.engine.list_session_ids_db(), [])
            self.assertFalse(self.engine.delete_session_db(self.sid))
            # صف غير موجود بعد إنشاء الجدول → load يعيد False
            self.assertTrue(self.engine.save_session_db(self.sid))
            fresh = UserTestingEngine()
            fresh.create_session("آخر")
            self.assertFalse(fresh.load_session_db("zz"))
        finally:
            config.DATABASE_PATH = original
            db_connection.close_pool()
            try:
                os.remove(tmp_db.name)
                os.remove(tmp_db.name + "-wal")
                os.remove(tmp_db.name + "-shm")
            except OSError:
                pass

    def test_db_exception_branches(self):
        with mock.patch("modules.user_testing.get_connection",
                        side_effect=RuntimeError("db down")):
            self.assertFalse(self.engine.save_session_db(self.sid))
            self.assertFalse(self.engine.load_session_db(self.sid))
            self.assertEqual(self.engine.list_session_ids_db(), [])
            self.assertFalse(self.engine.delete_session_db(self.sid))

    # ---------- تصدير مسارات الفشل ----------

    def test_export_excel_bad_path(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "no_such_dir", "x.xlsx")
            self.assertFalse(self.engine.export_excel(path))

    def test_export_csv_bad_path(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "no_such_dir", "x.csv")
            self.assertFalse(self.engine.export_csv(path))

    def test_export_csv_missing_session_id(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "x.csv")
            self.assertTrue(self.engine.export_csv(path, session_id="999"))

    @unittest.skipUnless(HAS_FPDF, "fpdf not installed")
    def test_export_pdf_no_arabic_font(self):
        with mock.patch.object(self.engine, "_arabic_font_path", return_value=None):
            with tempfile.TemporaryDirectory() as tmp:
                path = os.path.join(tmp, "x.pdf")
                self.assertTrue(self.engine.export_pdf(path))

    @unittest.skipUnless(HAS_FPDF, "fpdf not installed")
    def test_export_pdf_bad_path(self):
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "no_such_dir", "x.pdf")
            self.assertFalse(self.engine.export_pdf(path))

    def test_export_pdf_no_fpdf(self):
        with mock.patch.dict("sys.modules", {"fpdf": None}):
            with tempfile.TemporaryDirectory() as tmp:
                path = os.path.join(tmp, "x.pdf")
                self.assertFalse(self.engine.export_pdf(path))


class _FakePdf:
    def __init__(self, word_width):
        self._word_width = word_width

    def get_string_width(self, text):
        return sum(self._word_width for _ in text.split())


class UserTestingIntegrationTest(unittest.TestCase):
    """رحلة كاملة: كل المجموعات × كل السيناريوهات × كل التصنيفات."""

    def test_full_user_testing_journey(self):
        engine = UserTestingEngine()
        scores = {}
        for group in USER_GROUPS:
            session = engine.create_session(
                name=f"اختبار {group}", tester_name=group,
                user_group=group, scenario=SCENARIOS[0],
                environment="Windows 11",
            )
            for i, scenario in enumerate(SCENARIOS):
                category = FEEDBACK_CATEGORIES[i % len(FEEDBACK_CATEGORIES)]
                rating = 3 + (i % 3)
                engine.add_feedback(
                    session["id"], category=category,
                    comment=f"{group}/{scenario}", rating=rating,
                    scenario=scenario,
                )
            scores[group] = engine.satisfaction_score(session["id"])

        for group in USER_GROUPS:
            self.assertEqual(scores[group]["count"], len(SCENARIOS))

        # تقرير + قوائم
        report = engine.feedback_report("1")
        self.assertEqual(report["total_feedback"], len(SCENARIOS))
        self.assertGreaterEqual(len(engine.issue_list("1")), 0)
        self.assertGreaterEqual(len(engine.enhancement_requests("1")), 0)

        # JSON تصدير/استيراد
        with tempfile.TemporaryDirectory() as tmp:
            path = os.path.join(tmp, "all.json")
            self.assertTrue(engine.export_json(path))
            fresh = UserTestingEngine()
            imported = fresh.import_json(path)
            self.assertEqual(imported, len(USER_GROUPS))
            self.assertEqual(
                fresh.satisfaction_score("1")["overall"],
                engine.satisfaction_score("1")["overall"],
            )

        # DB حفظ واسترجاع
        original = config.DATABASE_PATH
        tmp_db = tempfile.NamedTemporaryFile(suffix=".db", delete=False)
        tmp_db.close()
        try:
            config.DATABASE_PATH = tmp_db.name
            db_connection.close_pool()
            for session in engine.list_sessions():
                self.assertTrue(engine.save_session_db(session["id"]))
            self.assertEqual(len(engine.list_session_ids_db()), len(USER_GROUPS))
            fresh2 = UserTestingEngine()
            for session in engine.list_sessions():
                self.assertTrue(fresh2.load_session_db(session["id"]))
            self.assertEqual(fresh2.satisfaction_score("1")["overall"],
                             engine.satisfaction_score("1")["overall"])
        finally:
            config.DATABASE_PATH = original
            db_connection.close_pool()
            try:
                os.remove(tmp_db.name)
                os.remove(tmp_db.name + "-wal")
                os.remove(tmp_db.name + "-shm")
            except OSError:
                pass


if __name__ == "__main__":
    unittest.main()
