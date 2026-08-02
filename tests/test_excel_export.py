# Unit tests for modules/excel_export.py (ExcelExporter).
# Covers the optional report sheets, fallback/exception branches,
# the openpyxl ImportError path and the export_comparison method.

import builtins
import os
import sys
import tempfile
import unittest
from unittest import mock

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import openpyxl.chart

from modules.excel_export import ExcelExporter


class TestExcelExporter(unittest.TestCase):
    """Tests for the advanced multi-sheet Excel exporter."""

    def setUp(self):
        self.exporter = ExcelExporter()
        self._tmp = tempfile.TemporaryDirectory()
        self.addCleanup(self._tmp.cleanup)

    def _path(self, name):
        return os.path.join(self._tmp.name, name)

    def _full_data(self):
        return {
            "cash": 1000,
            "accounts_receivable": 2000,
            "inventory": 3000,
            "current_assets": 6000,
            "fixed_assets": 4000,
            "intangible_assets": 1000,
            "total_assets": 11000,
            "accounts_payable": 2000,
            "short_term_debt": 1000,
            "current_liabilities": 3000,
            "long_term_debt": 2000,
            "long_term_liabilities": 2000,
            "share_capital": 5000,
            "retained_earnings": 1000,
            "equity": 6000,
            "total_liabilities": 5000,
            "revenue": 20000,
            "cost_of_goods_sold": 12000,
            "gross_profit": 8000,
            "operating_expenses": 4000,
            "net_income": 4000,
            "salaries_expense": 2000,
            "rent_expense": 1000,
            "utilities_expense": 500,
            "depreciation": 500,
            "interest_expense": 200,
            "income_tax": 800,
        }

    def test_export_full_report_all_sections(self):
        """Export a full report exercising every optional sheet."""
        ratios = {
            "current_ratio": 2.0, "quick_ratio": 1.5, "cash_ratio": 0.5,
            "gross_profit_margin": 40.0, "net_profit_margin": 20.0,
            "roa": 10.0, "roe": 15.0, "debt_to_equity": 0.5,
            "debt_ratio": 0.3, "interest_coverage": 5.0,
            "asset_turnover": 1.5, "inventory_turnover": 4.0,
            "receivable_turnover": 6.0,
        }
        tax_data = {"ibs": 1000, "tva": 2000, "irg": 500, "cnas": 300,
                    "cnac": 100, "versement_forfaitaire": 50}
        cashflow_data = {"net_income": 4000, "depreciation": 500,
                         "working_capital": -300, "operating": 4200,
                         "capex": -2000, "asset_sales": 0, "investing": -2000,
                         "debt": 1000, "dividends": -500, "financing": 500,
                         "net_change": 2700}
        budget_data = {"items": [
            {"name": "Revenue", "budget": 20000, "actual": 21000},
            {"name": "COGS", "budget": 12000, "actual": 11000},
        ]}
        cost_centers = [{"name": "Sales", "budget": 1000, "actual": 900,
                         "efficiency": 90.0}]
        comparative_data = {"years": ["2023", "2024"],
                            "metrics": {"ROE": [12.0, 15.0], "ROA": [8.0, 10.0]}}
        path = self._path("full_report.xlsx")
        result = self.exporter.export_full_report(
            path, self._full_data(), company_name="Company A",
            fiscal_year="2024", ratios=ratios, tax_data=tax_data,
            cashflow_data=cashflow_data, budget_data=budget_data,
            cost_centers=cost_centers, comparative_data=comparative_data,
        )
        self.assertTrue(result)
        self.assertTrue(os.path.exists(path))
        self.assertGreater(os.path.getsize(path), 0)
        wb = openpyxl.load_workbook(path)
        self.assertEqual(wb.sheetnames, [
            "Cover", "Balance Sheet", "Income Statement", "Financial Ratios",
            "Tax Data", "Cash Flow", "Budget vs Actual", "Cost Centers",
            "Comparative Analysis",
        ])

    def test_export_full_report_cover_only(self):
        """Export with only the core financial data."""
        path = self._path("cover_only.xlsx")
        result = self.exporter.export_full_report(
            path, self._full_data(), company_name="Company B", fiscal_year="2025"
        )
        self.assertTrue(result)
        wb = openpyxl.load_workbook(path)
        self.assertEqual(wb.sheetnames, ["Cover", "Balance Sheet", "Income Statement"])

    def test_export_full_report_budget_default_items(self):
        """Budget sheet falls back to default items when none supplied."""
        path = self._path("budget_default.xlsx")
        result = self.exporter.export_full_report(
            path, self._full_data(), budget_data={"note": "no items"}
        )
        self.assertTrue(result)
        wb = openpyxl.load_workbook(path)
        self.assertIn("Budget vs Actual", wb.sheetnames)

    def test_export_full_report_save_error_returns_false(self):
        """A workbook save failure is reported as False."""
        path = self._path("fail.xlsx")
        with mock.patch.object(openpyxl.Workbook, "save",
                               side_effect=PermissionError("denied")):
            result = self.exporter.export_full_report(path, self._full_data())
        self.assertFalse(result)
        self.assertFalse(os.path.exists(path))

    def test_income_chart_error_is_swallowed(self):
        """Chart creation errors must not abort the export."""
        path = self._path("no_chart.xlsx")
        with mock.patch.object(openpyxl.chart, "PieChart",
                               side_effect=RuntimeError("boom")):
            result = self.exporter.export_full_report(path, self._full_data())
        self.assertTrue(result)
        self.assertTrue(os.path.exists(path))

    def test_init_styles_import_error_fallback(self):
        """_init_styles must not crash when openpyxl is missing."""
        class _NoOpenpyxlExporter(ExcelExporter):
            TITLE_FONT = None

        real_import = builtins.__import__

        def _block_openpyxl(name, *args, **kwargs):
            if name.startswith("openpyxl"):
                raise ImportError("No module named 'openpyxl.styles'")
            return real_import(name, *args, **kwargs)

        with mock.patch.object(builtins, "__import__", side_effect=_block_openpyxl):
            _NoOpenpyxlExporter._init_styles()
        self.assertIsNone(_NoOpenpyxlExporter.TITLE_FONT)

    def test_remove_default_sheet_deletes_active(self):
        """_remove_default_sheet removes the leftover Sheet tab."""
        wb = openpyxl.Workbook()
        self.assertIn("Sheet", wb.sheetnames)
        self.exporter._remove_default_sheet(wb)
        self.assertNotIn("Sheet", wb.sheetnames)

    def test_export_comparison_success(self):
        """Export the benchmark comparison report."""
        data = [
            {"Ratio": "Current Ratio", "Company Value": 2.0, "Sector Min": 1.0,
             "Sector Avg": 1.5, "Sector Max": 3.0, "Status": "Good", "Score": 80},
            {"Ratio": "ROE", "Company Value": 12.0, "Sector Min": 5.0,
             "Sector Avg": 10.0, "Sector Max": 18.0, "Status": "Excellent",
             "Score": 90},
        ]
        path = self._path("benchmark.xlsx")
        result = self.exporter.export_comparison(
            data, path, title="Benchmark Comparison"
        )
        self.assertTrue(result)
        self.assertTrue(os.path.exists(path))
        self.assertGreater(os.path.getsize(path), 0)
        wb = openpyxl.load_workbook(path)
        self.assertEqual(wb.sheetnames, ["Benchmark"])

    def test_export_comparison_save_error_returns_false(self):
        """A benchmark save failure is reported as False."""
        path = self._path("bench_fail.xlsx")
        with mock.patch.object(openpyxl.Workbook, "save",
                               side_effect=PermissionError("denied")):
            result = self.exporter.export_comparison([], path)
        self.assertFalse(result)


if __name__ == "__main__":
    unittest.main()
