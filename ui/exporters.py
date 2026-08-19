# -*- coding: utf-8 -*-
"""طبقة التصدير الموحدة — أدوات مشتركة لتصدير PDF/Excel عبر الشاشات

تجمّع الأنماط المتكررة في كل الشاشات:
  - مربع الحفظ الموحد (QFileDialog)
  - تنسيق رأس أوراق Excel (openpyxl)
  - بناء مصنف Excel متعدد الأوراق (رأس ملوّن + بيانات)
  - حفظ أشكال matplotlib في PDF واحد (PdfPages)
"""
from PyQt6.QtWidgets import QFileDialog
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2980B9", end_color="2980B9",
                          fill_type="solid")
DEFAULT_SHEET_FILL = "1F4E79"


def ask_save_path(parent, caption, default_name, file_filter):
    """مربع حفظ موحد — يعيد المسار المختار أو None عند الإلغاء."""
    path, _ = QFileDialog.getSaveFileName(parent, caption, default_name,
                                          file_filter)
    return path or None


def style_header_row(ws, row=1):
    """تنسيق موحد لصف الرأس (خط أبيض عريض على خلفية زرقاء)."""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL


def new_workbook():
    """مصنف Excel جديد."""
    return Workbook()


def add_excel_sheet(wb, title, headers, rows, header_fill=DEFAULT_SHEET_FILL):
    """إضافة ورقة بعنوان + صف رأس ملوّن + صفوف بيانات.

    - ``title``: اسم الورقة (يستعمل ورقة المصنف النشطة إذا كانت بلا اسم).
    - ``rows``: تكرارات من القيم تُضاف سطراً بسطر.
    - ``header_fill``: لون خلفية صف الرأس (hex مثل "1F4E79").
    يُرجع الورقة للاستخدامات الإضافية.
    """
    if len(wb.sheetnames) == 1 and wb.active.title == "Sheet":
        ws = wb.active
    else:
        ws = wb.create_sheet(title)
    ws.title = title
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    fill = PatternFill(start_color=header_fill, end_color=header_fill,
                       fill_type="solid")
    for cell in ws[1]:
        cell.font = HEADER_FONT
        cell.fill = fill
    return ws


def write_charts_pdf(path, widgets):
    """حفظ عدة أشكال pyqtgraph في ملف PDF واحد."""
    from PyQt6.QtGui import QPdfWriter, QPainter
    from PyQt6.QtCore import QRect
    if not widgets:
        return
    writer = QPdfWriter(path)
    writer.setResolution(150)
    painter = QPainter()
    painter.begin(writer)
    for i, w in enumerate(widgets):
        pixmap = w.grab()
        img = pixmap.toImage()
        painter.drawImage(QRect(0, 0, writer.width(), writer.height()), img)
        if i < len(widgets) - 1:
            writer.newPage()
    painter.end()
