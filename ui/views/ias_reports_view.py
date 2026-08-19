# واجهة التقارير المالية المعيارية IAS/IFRS
# =========================================
# IAS 1 (المركز المالي/الدخل/التغيرات في حقوق الملكية) + IAS 7 (التدفقات النقدية)

from ui.views._path import _  # noqa: F401

from PyQt6.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QTableWidget,
    QTableWidgetItem, QTabWidget, QMessageBox, QFileDialog, QHeaderView,
)
from PyQt6.QtCore import Qt

from ui.views._base import BaseView
from ui.resources.i18n import t
from modules.ias_reports import generate_all


class IASReportsView(BaseView):
    """تقارير IAS/IFRS المعيارية"""

    def __init__(self):
        super().__init__()
        self._result = {}
        self.setup_ui()
        self.refresh()

    def setup_ui(self):
        self._make_header("ias_title", "ias_subtitle")

        self.tabs = QTabWidget()

        self._bs_table = QTableWidget()
        self._is_table = QTableWidget()
        self._cf_table = QTableWidget()
        self._eq_table = QTableWidget()

        for tbl in (self._bs_table, self._is_table, self._cf_table, self._eq_table):
            tbl.setEditTriggers(QTableWidget.EditTrigger.NoEditTriggers)
            tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
            tbl.setSelectionBehavior(QTableWidget.SelectionBehavior.SelectRows)

        self.tabs.addTab(self._bs_table, t("ias_tab_balance_sheet"))
        self.tabs.addTab(self._is_table, t("ias_tab_income"))
        self.tabs.addTab(self._cf_table, t("ias_tab_cash_flow"))
        self.tabs.addTab(self._eq_table, t("ias_tab_equity"))
        self._main_layout.addWidget(self.tabs)

        btns = QHBoxLayout()
        self.refresh_btn = QPushButton(t("ias_refresh"))
        self.refresh_btn.clicked.connect(self.refresh)
        btns.addWidget(self.refresh_btn)
        self.pdf_btn = QPushButton(t("ias_export_pdf"))
        self.pdf_btn.clicked.connect(self._export_pdf)
        btns.addWidget(self.pdf_btn)
        self.excel_btn = QPushButton(t("ias_export_excel"))
        self.excel_btn.clicked.connect(self._export_excel)
        btns.addWidget(self.excel_btn)
        btns.addStretch()
        self._main_layout.addLayout(btns)

        self._main_layout.addStretch()

    # ── Refresh ──────────────────────────────────────────────────────────────

    def refresh(self):
        self._result = generate_all()
        self._render_balance_sheet(self._result.get("balance_sheet", {}))
        self._render_income_statement(self._result.get("income_statement", {}))
        self._render_cash_flow(self._result.get("cash_flow", {}))
        self._render_equity(self._result.get("equity_statement", {}))

    def _set_header(self, table, headers):
        table.setColumnCount(len(headers))
        table.setHorizontalHeaderLabels(headers)

    def _add_row(self, table, row, label, amount, bold=False):
        lbl = QTableWidgetItem(label)
        val = QTableWidgetItem(f"{amount:,.2f}")
        val.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
        if bold:
            from PyQt6.QtGui import QFont
            f = QFont()
            f.setBold(True)
            lbl.setFont(f)
            val.setFont(f)
        table.setItem(row, 0, lbl)
        table.setItem(row, 1, val)

    # ── Balance Sheet ────────────────────────────────────────────────────────

    def _render_balance_sheet(self, bs):
        table = self._bs_table
        table.clear()
        comp = self._result.get("company_name", "")
        fy = self._result.get("fiscal_year", "")
        title = t("ias_bs_title").format(company=comp, year=fy)
        self._set_header(table, [title, t("ias_amount")])

        assets = bs.get("assets", {})
        eqliab = bs.get("equity_liabilities", {})

        rows = []
        rows.append(("H", t("ias_assets")))
        rows.append(("SH", t("ias_non_current_assets")))
        for label_key, amt in assets.get("non_current", []):
            rows.append(("R", (t(label_key), amt)))
        rows.append(("T", (t("ias_total_non_current"), assets.get("total_non_current", 0))))
        rows.append(("SH", t("ias_current_assets")))
        for label_key, amt in assets.get("current", []):
            rows.append(("R", (t(label_key), amt)))
        rows.append(("T", (t("ias_total_current"), assets.get("total_current", 0))))
        rows.append(("GT", (t("ias_total_assets"), bs.get("total_assets", 0))))

        rows.append(("H", t("ias_equity_liabilities")))
        rows.append(("SH", t("ias_equity")))
        for label_key, amt in eqliab.get("equity", []):
            rows.append(("R", (t(label_key), amt)))
        rows.append(("T", (t("ias_total_equity"), eqliab.get("total_equity", 0))))
        rows.append(("SH", t("ias_non_current_liabilities")))
        for label_key, amt in eqliab.get("non_current_liabilities", []):
            if amt > 0:
                rows.append(("R", (t(label_key), amt)))
        rows.append(("T", (t("ias_total_non_current_liabilities"), eqliab.get("total_non_current_liabilities", 0))))
        rows.append(("SH", t("ias_current_liabilities")))
        for label_key, amt in eqliab.get("current_liabilities", []):
            if amt > 0:
                rows.append(("R", (t(label_key), amt)))
        rows.append(("T", (t("ias_total_current_liabilities"), eqliab.get("total_current_liabilities", 0))))
        rows.append(("GT", (t("ias_total_equity_liabilities"), bs.get("total_equity_liabilities", 0))))

        table.setRowCount(len(rows))
        for i, (rtype, content) in enumerate(rows):
            if rtype == "H":
                lbl = QTableWidgetItem(content)
                from PyQt6.QtGui import QFont
                f = QFont()
                f.setBold(True)
                f.setPointSize(12)
                lbl.setFont(f)
                table.setItem(i, 0, lbl)
                table.setItem(i, 1, QTableWidgetItem(""))
            elif rtype == "SH":
                lbl = QTableWidgetItem(content)
                from PyQt6.QtGui import QFont
                f = QFont()
                f.setBold(True)
                f.setPointSize(10)
                lbl.setFont(f)
                table.setItem(i, 0, lbl)
                table.setItem(i, 1, QTableWidgetItem(""))
            elif rtype in ("T", "GT"):
                label, amount = content
                self._add_row(table, i, label, amount, bold=True)
            else:
                label, amount = content
                self._add_row(table, i, label, amount)

    # ── Income Statement ──────────────────────────────────────────────────────

    def _render_income_statement(self, inc):
        table = self._is_table
        table.clear()
        comp = self._result.get("company_name", "")
        fy = self._result.get("fiscal_year", "")
        self._set_header(table, [t("ias_is_title").format(company=comp, year=fy), t("ias_amount")])

        rows = []
        for label_key, amt in inc.get("items", []):
            rows.append(("R", (t(label_key), amt)))
        rows.append(("T", (t("ias_gross_profit"), inc.get("gross_profit", 0))))
        for label_key, amt in inc.get("operating_items", []):
            rows.append(("R", (t(label_key), amt)))
        rows.append(("T", (t("ias_operating_profit"), inc.get("operating_profit", 0))))
        rows.append(("R", (t("ias_profit_before_tax"), inc.get("profit_before_tax", 0))))
        rows.append(("R", (t("ias_tax_expense"), inc.get("tax_expense", 0))))
        rows.append(("GT", (t("ias_net_income"), inc.get("net_income", 0))))

        table.setRowCount(len(rows))
        for i, (rtype, content) in enumerate(rows):
            label, amount = content
            self._add_row(table, i, label, amount, bold=(rtype in ("T", "GT")))

    # ── Cash Flow ────────────────────────────────────────────────────────────

    def _render_cash_flow(self, cf):
        table = self._cf_table
        table.clear()
        comp = self._result.get("company_name", "")
        fy = self._result.get("fiscal_year", "")
        self._set_header(table, [t("ias_cf_title").format(company=comp, year=fy), t("ias_amount")])

        rows = []
        rows.append(("H", t("ias_operating_activities")))
        for label_key, amt in cf.get("operating", []):
            rows.append(("R", (t(label_key), amt)))
        rows.append(("T", (t("ias_net_operating"), cf.get("operating_total", 0))))
        rows.append(("H", t("ias_investing_activities")))
        for label_key, amt in cf.get("investing", []):
            rows.append(("R", (t(label_key), amt)))
        rows.append(("T", (t("ias_net_investing"), cf.get("investing_total", 0))))
        rows.append(("H", t("ias_financing_activities")))
        for label_key, amt in cf.get("financing", []):
            rows.append(("R", (t(label_key), amt)))
        rows.append(("T", (t("ias_net_financing"), cf.get("financing_total", 0))))
        rows.append(("GT", (t("ias_net_cash_change"), cf.get("net_change", 0))))
        rows.append(("R", (t("ias_cash_beginning"), cf.get("cash_beginning", 0))))
        rows.append(("GT", (t("ias_cash_ending"), cf.get("cash_ending", 0))))

        table.setRowCount(len(rows))
        for i, (rtype, content) in enumerate(rows):
            if rtype == "H":
                lbl = QTableWidgetItem(content)
                from PyQt6.QtGui import QFont
                f = QFont()
                f.setBold(True)
                f.setPointSize(11)
                lbl.setFont(f)
                table.setItem(i, 0, lbl)
                table.setItem(i, 1, QTableWidgetItem(""))
            else:
                label, amount = content
                self._add_row(table, i, label, amount, bold=(rtype in ("T", "GT")))

    # ── Equity Statement ──────────────────────────────────────────────────────

    def _render_equity(self, eq_rpt):
        table = self._eq_table
        table.clear()
        comp = self._result.get("company_name", "")
        fy = self._result.get("fiscal_year", "")
        self._set_header(table, [t("ias_eq_title").format(company=comp, year=fy), t("ias_amount")])

        rows = []
        rows.append(("R", (t("ias_opening_equity"), eq_rpt.get("opening_balance", 0))))
        for label_key, amt in eq_rpt.get("changes", []):
            rows.append(("R", (t(label_key), amt)))
        rows.append(("GT", (t("ias_closing_equity"), eq_rpt.get("closing_balance", 0))))

        table.setRowCount(len(rows))
        for i, (rtype, content) in enumerate(rows):
            label, amount = content
            self._add_row(table, i, label, amount, bold=(rtype == "GT"))

    # ── Export ───────────────────────────────────────────────────────────────

    def _export_pdf(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, t("ias_export_pdf"), "ias_reports.pdf", "PDF Files (*.pdf)"
        )
        if not file_path:
            return
        try:
            from fpdf import FPDF
            pdf = FPDF()
            pdf.add_page()
            pdf.add_font("Amiri", "", "fonts/Amiri-Regular.ttf", uni=True)
            pdf.add_font("Amiri", "B", "fonts/Amiri-Bold.ttf", uni=True)
            pdf.set_font("Amiri", "B", 14)
            comp = self._result.get("company_name", "") or self._result.get("company_name_fr", "")
            pdf.cell(0, 10, comp or t("ias_title"), ln=True, align="C")
            pdf.set_font("Amiri", "", 10)
            for tab_title, data_key in (
                (t("ias_tab_balance_sheet"), "balance_sheet"),
                (t("ias_tab_income"), "income_statement"),
                (t("ias_tab_cash_flow"), "cash_flow"),
                (t("ias_tab_equity"), "equity_statement"),
            ):
                pdf.add_page()
                pdf.set_font("Amiri", "B", 12)
                pdf.cell(0, 8, tab_title, ln=True)
                pdf.set_font("Amiri", "", 9)
                for tbl in self._export_tables_for_pdf(data_key):
                    for row in tbl:
                        pdf.cell(95, 6, str(row[0]), border=1)
                        pdf.cell(95, 6, str(row[1]), border=1, ln=True)
                pdf.ln(4)
            pdf.output(file_path)
            QMessageBox.information(self, t("success"), f"✅ {file_path}")
        except Exception as e:
            QMessageBox.critical(self, t("error"), str(e))

    def _export_tables_for_pdf(self, key):
        data = self._result.get(key, {})
        tables = []
        if key == "balance_sheet":
            tables.append(
                [("— " + t("ias_bs_title").format(company="", year=""), "—")]
                + [
                    (t("ias_assets"), ""), (t("ias_total_assets"),
                    f"{data.get('total_assets', 0):,.2f}")
                ]
            )
        elif key == "income_statement":
            tables.append(
                [(t(lk), f"{a:,.2f}") for lk, a in data.get("items", [])]
                + [(t("ias_net_income"), f"{data.get('net_income', 0):,.2f}")]
            )
        elif key == "cash_flow":
            tables.append(
                [(t(lk), f"{a:,.2f}") for lk, a in data.get("operating", [])]
                + [(t("ias_net_cash_change"), f"{data.get('net_change', 0):,.2f}")]
            )
        elif key == "equity_statement":
            tables.append([
                (t("ias_opening_equity"), f"{data.get('opening_balance', 0):,.2f}"),
                (t("ias_closing_equity"), f"{data.get('closing_balance', 0):,.2f}"),
            ])
        return tables

    def _export_excel(self):
        file_path, _ = QFileDialog.getSaveFileName(
            self, t("ias_export_excel"), "ias_reports.xlsx", "Excel Files (*.xlsx)"
        )
        if not file_path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            wb = Workbook()
            thin = Side(style="thin")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)

            def _sheet(name, rows_func):
                ws = wb.create_sheet(title=name[:31])
                ws.append([t("ias_description"), t("ias_amount")])
                for label, amount in rows_func():
                    ws.append([label, amount])
                for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
                    for cell in row:
                        cell.border = border
                ws.column_dimensions["A"].width = 45
                ws.column_dimensions["B"].width = 20

            data = self._result
            bs = data.get("balance_sheet", {})
            assets = bs.get("assets", {})
            eqliab = bs.get("equity_liabilities", {})

            def bs_rows():
                r = [(t("ias_assets"), "")]
                for lk, a in assets.get("non_current", []): r.append((t(lk), a))
                r.append((t("ias_total_non_current"), assets.get("total_non_current", 0)))
                for lk, a in assets.get("current", []): r.append((t(lk), a))
                r.append((t("ias_total_current"), assets.get("total_current", 0)))
                r.append((t("ias_total_assets"), bs.get("total_assets", 0)))
                r.append((t("ias_equity_liabilities"), ""))
                for lk, a in eqliab.get("equity", []): r.append((t(lk), a))
                r.append((t("ias_total_equity"), eqliab.get("total_equity", 0)))
                for lk, a in eqliab.get("current_liabilities", []):
                    if a > 0: r.append((t(lk), a))
                r.append((t("ias_total_current_liabilities"), eqliab.get("total_current_liabilities", 0)))
                r.append((t("ias_total_equity_liabilities"), bs.get("total_equity_liabilities", 0)))
                return r

            _sheet(t("ias_tab_balance_sheet"), bs_rows)

            inc = data.get("income_statement", {})

            def is_rows():
                r = []
                for lk, a in inc.get("items", []): r.append((t(lk), a))
                r.append((t("ias_gross_profit"), inc.get("gross_profit", 0)))
                for lk, a in inc.get("operating_items", []): r.append((t(lk), a))
                r.append((t("ias_operating_profit"), inc.get("operating_profit", 0)))
                r.append((t("ias_profit_before_tax"), inc.get("profit_before_tax", 0)))
                r.append((t("ias_tax_expense"), inc.get("tax_expense", 0)))
                r.append((t("ias_net_income"), inc.get("net_income", 0)))
                return r

            _sheet(t("ias_tab_income"), is_rows)

            cf = data.get("cash_flow", {})

            def cf_rows():
                r = [(t("ias_operating_activities"), "")]
                for lk, a in cf.get("operating", []): r.append((t(lk), a))
                r.append((t("ias_net_operating"), cf.get("operating_total", 0)))
                r.append((t("ias_net_cash_change"), cf.get("net_change", 0)))
                r.append((t("ias_cash_ending"), cf.get("cash_ending", 0)))
                return r

            _sheet(t("ias_tab_cash_flow"), cf_rows)

            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]
            wb.save(file_path)
            QMessageBox.information(self, t("success"), f"✅ {file_path}")
        except Exception as e:
            QMessageBox.critical(self, t("error"), str(e))
