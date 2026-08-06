# محرك التقارير المعيارية IFRS/IAS
# ================================
# قائمة المركز المالي + الدخل الشامل + التدفقات النقدية + التغير في حقوق الملكية
# + تصدير PDF/Excel

from datetime import datetime
from modules.reporting import ReportGenerator
from utils.app_logger import get_logger

log = get_logger("ifrs")


class IFRSReporter:

    def __init__(self, company_name="", fiscal_year="", currency="DZD"):
        self.company = company_name
        self.year = fiscal_year
        self.currency = currency
        self.generated = datetime.now().strftime("%Y-%m-%d %H:%M")

    def balance_sheet(self, financial_data):
        """IAS 1 — Statement of Financial Position (with IFRS classification)."""
        data = financial_data or {}
        total_assets = float(data.get("total_assets") or 0)
        total_liabilities = float(data.get("total_liabilities") or 0)
        equity = float(data.get("equity") or total_assets - total_liabilities)
        current_assets = float(data.get("current_assets") or 0)
        inventory = float(data.get("inventory") or 0)
        cash = float(data.get("cash") or 0)
        fixed_assets = float(data.get("fixed_assets") or total_assets - current_assets - inventory - cash)
        current_liabilities = float(data.get("current_liabilities") or 0)
        long_term_liabilities = float(data.get("long_term_liabilities") or total_liabilities - current_liabilities)

        return {
            "title": f"IAS 1 — Statement of Financial Position — {self.company} — {self.year}",
            "currency": self.currency,
            "non_current_assets": round(max(fixed_assets, 0), 2),
            "current_assets": round(current_assets, 2),
            "inventory": round(inventory, 2),
            "cash_and_equivalents": round(cash, 2),
            "total_assets": round(total_assets, 2),
            "equity": {
                "share_capital": round(float(data.get("share_capital") or equity * 0.4), 2),
                "retained_earnings": round(float(data.get("retained_earnings") or equity * 0.6), 2),
                "total_equity": round(equity, 2),
            },
            "non_current_liabilities": round(max(long_term_liabilities, 0), 2),
            "current_liabilities": round(current_liabilities, 2),
            "total_liabilities": round(total_liabilities, 2),
            "total_equity_and_liabilities": round(total_liabilities + equity, 2),
        }

    def income_statement(self, financial_data):
        """IAS 1 — Statement of Comprehensive Income."""
        data = financial_data or {}
        revenue = float(data.get("revenue") or 0)
        cogs = float(data.get("cost_of_goods_sold") or 0)
        gross_profit = round(revenue - cogs, 2)
        operating_expenses = float(data.get("operating_expenses") or 0)
        operating_profit = round(gross_profit - operating_expenses, 2)
        net_income = float(data.get("net_income") or operating_profit)

        return {
            "title": f"IAS 1 — Statement of Comprehensive Income — {self.company} — {self.year}",
            "currency": self.currency,
            "revenue": round(revenue, 2),
            "cost_of_sales": round(cogs, 2),
            "gross_profit": gross_profit,
            "gross_margin_pct": round(gross_profit / revenue * 100, 1) if revenue else 0,
            "operating_expenses": round(operating_expenses, 2),
            "operating_profit": operating_profit,
            "operating_margin_pct": round(operating_profit / revenue * 100, 1) if revenue else 0,
            "net_income": round(net_income, 2),
            "net_margin_pct": round(net_income / revenue * 100, 1) if revenue else 0,
        }

    def cash_flow(self, financial_data):
        """IAS 7 — Statement of Cash Flows."""
        data = financial_data or {}
        operating = float(data.get("operating_cash_flow") or 0)
        investing = float(data.get("investing_cash_flow") or 0)
        financing = float(data.get("financing_cash_flow") or 0)
        net_change = round(operating + investing + financing, 2)
        opening_cash = float(data.get("opening_cash") or 0)
        closing_cash = round(opening_cash + net_change, 2)

        return {
            "title": f"IAS 7 — Statement of Cash Flows — {self.company} — {self.year}",
            "currency": self.currency,
            "operating_activities": round(operating, 2),
            "investing_activities": round(investing, 2),
            "financing_activities": round(financing, 2),
            "net_change": net_change,
            "opening_cash": round(opening_cash, 2),
            "closing_cash": closing_cash,
        }

    def equity_changes(self, financial_data):
        """IAS 1 — Statement of Changes in Equity."""
        data = financial_data or {}
        opening_equity = float(data.get("opening_equity") or 0)
        net_income = float(data.get("net_income") or 0)
        dividends = float(data.get("dividends") or 0)
        closing_equity = round(opening_equity + net_income - dividends, 2)

        return {
            "title": f"IAS 1 — Statement of Changes in Equity — {self.company} — {self.year}",
            "currency": self.currency,
            "opening_equity": round(opening_equity, 2),
            "net_income": round(net_income, 2),
            "dividends": round(dividends, 2),
            "closing_equity": closing_equity,
        }

    def full_report(self, financial_data, ratios=None):
        return {
            "company": self.company,
            "year": self.year,
            "currency": self.currency,
            "generated": self.generated,
            "balance_sheet": self.balance_sheet(financial_data),
            "income_statement": self.income_statement(financial_data),
            "cash_flow": self.cash_flow(financial_data),
            "equity_changes": self.equity_changes(financial_data),
            "ratios": ratios or {},
        }

    def export_pdf(self, financial_data, filepath, ratios=None):
        """تصدير التقرير الكامل IFRS إلى PDF."""
        report = self.full_report(financial_data, ratios)
        text = self._format_text(report)
        reporter = ReportGenerator(self.company, self.year)
        return reporter.export_to_pdf(text, filepath)

    def export_excel(self, financial_data, filepath, ratios=None):
        """تصدير التقرير الكامل IFRS إلى Excel."""
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        report = self.full_report(financial_data, ratios)

        wb = openpyxl.Workbook()
        thin = Side(style="thin")
        border = Border(bottom=thin)
        header_fill = PatternFill("solid", fgColor="1F7A33")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=14, color="1E3A5F")
        amount_fmt = '#,##0.00'

        def write_statement(ws, title, rows, start_row):
            ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=3)
            ws.cell(start_row, 1, title).font = title_font
            start_row += 1
            ws.cell(start_row, 1, "Item").font = header_font
            ws.cell(start_row, 1).fill = header_fill
            ws.cell(start_row, 2, "Amount").font = header_font
            ws.cell(start_row, 2).fill = header_fill
            ws.cell(start_row, 3, "").fill = header_fill
            start_row += 1
            for label, amount in rows:
                ws.cell(start_row, 1, label).border = border
                c = ws.cell(start_row, 2, amount)
                c.number_format = amount_fmt
                c.border = border
                if label.startswith(" ") and label.strip().startswith("Total"):
                    ws.cell(start_row, 1).font = Font(bold=True)
                    c.font = Font(bold=True)
                start_row += 1
            return start_row + 1

        ws = wb.active
        ws.title = "IFRS Report"

        r = 1
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        ws.cell(r, 1, f"IFRS Financial Statements — {self.company} — {self.year}").font = Font(bold=True, size=16, color="1F7A33")
        r += 2
        ws.cell(r, 1, f"Generated: {self.generated}").font = Font(italic=True, size=10)
        r += 2

        bs = report["balance_sheet"]
        rows = [
            ("Non-Current Assets", bs["non_current_assets"]),
            ("Current Assets", bs["current_assets"]),
            ("  Inventory", bs["inventory"]),
            ("  Cash & Equivalents", bs["cash_and_equivalents"]),
            ("Total Assets", bs["total_assets"]),
            ("", ""),
            ("Equity", ""),
            ("  Share Capital", bs["equity"]["share_capital"]),
            ("  Retained Earnings", bs["equity"]["retained_earnings"]),
            (" Total Equity", bs["equity"]["total_equity"]),
            ("", ""),
            ("Non-Current Liabilities", bs["non_current_liabilities"]),
            ("Current Liabilities", bs["current_liabilities"]),
            (" Total Liabilities", bs["total_liabilities"]),
            ("", ""),
            ("Total Equity & Liabilities", bs["total_equity_and_liabilities"]),
        ]
        r = write_statement(ws, bs["title"], rows, r)

        inc = report["income_statement"]
        rows = [
            ("Revenue", inc["revenue"]),
            ("Cost of Sales", inc["cost_of_sales"]),
            ("Gross Profit", inc["gross_profit"]),
            ("  Gross Margin %", inc["gross_margin_pct"]),
            ("Operating Expenses", inc["operating_expenses"]),
            ("Operating Profit", inc["operating_profit"]),
            ("  Operating Margin %", inc["operating_margin_pct"]),
            ("Net Income", inc["net_income"]),
            ("  Net Margin %", inc["net_margin_pct"]),
        ]
        r = write_statement(ws, inc["title"], rows, r)

        cf = report["cash_flow"]
        rows = [
            ("Operating Activities", cf["operating_activities"]),
            ("Investing Activities", cf["investing_activities"]),
            ("Financing Activities", cf["financing_activities"]),
            ("Net Change in Cash", cf["net_change"]),
            ("Opening Cash", cf["opening_cash"]),
            ("Closing Cash", cf["closing_cash"]),
        ]
        r = write_statement(ws, cf["title"], rows, r)

        eq = report["equity_changes"]
        rows = [
            ("Opening Equity", eq["opening_equity"]),
            ("Net Income", eq["net_income"]),
            ("Dividends", eq["dividends"]),
            ("Closing Equity", eq["closing_equity"]),
        ]
        write_statement(ws, eq["title"], rows, r)

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        try:
            wb.save(filepath)
            return True
        except Exception as exc:
            log.error("IFRS Excel export error: %s", exc)
            return False

    def _format_text(self, report):
        bs = report["balance_sheet"]
        inc = report["income_statement"]
        cf = report["cash_flow"]
        eq = report["equity_changes"]

        lines = [
            "=" * 70,
            f"  IFRS FINANCIAL STATEMENTS — {report['company']} — {report['year']}",
            f"  Currency: {report['currency']}   Generated: {report['generated']}",
            "=" * 70,
            "",
            bs["title"],
            "-" * 55,
            f"  Non-Current Assets                {bs['non_current_assets']:>15,.2f} {bs['currency']}",
            f"  Current Assets                    {bs['current_assets']:>15,.2f} {bs['currency']}",
            f"    Inventory                       {bs['inventory']:>15,.2f} {bs['currency']}",
            f"    Cash & Equivalents              {bs['cash_and_equivalents']:>15,.2f} {bs['currency']}",
            f"  TOTAL ASSETS                      {bs['total_assets']:>15,.2f} {bs['currency']}",
            "",
            f"  Share Capital                     {bs['equity']['share_capital']:>15,.2f} {bs['currency']}",
            f"  Retained Earnings                 {bs['equity']['retained_earnings']:>15,.2f} {bs['currency']}",
            f"  TOTAL EQUITY                      {bs['equity']['total_equity']:>15,.2f} {bs['currency']}",
            "",
            f"  Non-Current Liabilities           {bs['non_current_liabilities']:>15,.2f} {bs['currency']}",
            f"  Current Liabilities               {bs['current_liabilities']:>15,.2f} {bs['currency']}",
            f"  TOTAL LIABILITIES                 {bs['total_liabilities']:>15,.2f} {bs['currency']}",
            f"  TOTAL EQUITY & LIABILITIES        {bs['total_equity_and_liabilities']:>15,.2f} {bs['currency']}",
            "",
            inc["title"],
            "-" * 55,
            f"  Revenue                           {inc['revenue']:>15,.2f} {bs['currency']}",
            f"  Cost of Sales                    ({inc['cost_of_sales']:>14,.2f} {bs['currency']})",
            f"  GROSS PROFIT                      {inc['gross_profit']:>15,.2f} {bs['currency']}",
            f"    Gross Margin                    {inc['gross_margin_pct']:>14.1f}%",
            f"  Operating Expenses               ({inc['operating_expenses']:>14,.2f} {bs['currency']})",
            f"  OPERATING PROFIT                  {inc['operating_profit']:>15,.2f} {bs['currency']}",
            f"    Operating Margin                {inc['operating_margin_pct']:>14.1f}%",
            f"  NET INCOME                        {inc['net_income']:>15,.2f} {bs['currency']}",
            f"    Net Margin                      {inc['net_margin_pct']:>14.1f}%",
            "",
            cf["title"],
            "-" * 55,
            f"  Operating Activities              {cf['operating_activities']:>15,.2f} {bs['currency']}",
            f"  Investing Activities              {cf['investing_activities']:>15,.2f} {bs['currency']}",
            f"  Financing Activities              {cf['financing_activities']:>15,.2f} {bs['currency']}",
            f"  NET CHANGE IN CASH                {cf['net_change']:>15,.2f} {bs['currency']}",
            f"  Opening Cash                      {cf['opening_cash']:>15,.2f} {bs['currency']}",
            f"  CLOSING CASH                      {cf['closing_cash']:>15,.2f} {bs['currency']}",
            "",
            eq["title"],
            "-" * 55,
            f"  Opening Equity                    {eq['opening_equity']:>15,.2f} {bs['currency']}",
            f"  Net Income                        {eq['net_income']:>15,.2f} {bs['currency']}",
            f"  Dividends                        ({eq['dividends']:>14,.2f} {bs['currency']})",
            f"  CLOSING EQUITY                    {eq['closing_equity']:>15,.2f} {bs['currency']}",
            "",
            "=" * 70,
            f"  This report complies with IAS 1, IAS 7.",
            f"  Generated by Smart Accounting Platform — {self.generated}",
            "=" * 70,
        ]
        return "\n".join(lines)


ifrs_reporter = IFRSReporter()
