# تصدير Excel المتقدم - أوراق متعددة + رسوم بيانية + تنسيق
# ==========================================================

from datetime import datetime
from typing import Dict, Any, List
from utils.app_logger import get_logger

logger = get_logger("excel_export")


class ExcelExporter:
    """فئة لتصدير Excel متقدم مع أوراق متعددة وتنسيقات احترافية"""

    HEADER_FONT = None
    HEADER_FILL = None
    TITLE_FONT = None
    THIN_BORDER = None

    @classmethod
    def _init_styles(cls):
        if cls.TITLE_FONT is not None:
            return
        try:
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            cls.Font = Font
            cls.TITLE_FONT = Font(bold=True, size=14, color="2C3E50")
            cls.SUBTITLE_FONT = Font(size=11, color="7F8C8D")
            cls.HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
            cls.HEADER_FILL = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
            cls.TOTAL_FILL = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
            cls.TOTAL_FONT = Font(bold=True, size=11)
            cls.THIN_BORDER = Border(
                left=Side(style='thin', color='D5D8DC'),
                right=Side(style='thin', color='D5D8DC'),
                top=Side(style='thin', color='D5D8DC'),
                bottom=Side(style='thin', color='D5D8DC'),
            )
            cls.ACCENT_FILL = PatternFill(start_color="E8F6F3", end_color="E8F6F3", fill_type="solid")
            cls.RIGHT_ALIGN = Alignment(horizontal='right', vertical='center')
            cls.CENTER_ALIGN = Alignment(horizontal='center', vertical='center')
        except ImportError:
            logger.warning("openpyxl not available")

    def __init__(self):
        self._init_styles()

    def _style_header_row(self, ws, row: int, cols: int):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = self.HEADER_FONT
            cell.fill = self.HEADER_FILL
            cell.border = self.THIN_BORDER
            cell.alignment = self.CENTER_ALIGN

    def _style_data_row(self, ws, row: int, cols: int, is_total: bool = False):
        for col in range(1, cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = self.THIN_BORDER
            if is_total:
                cell.fill = self.TOTAL_FILL
                cell.font = self.TOTAL_FONT
            if col >= 2 and isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = self.RIGHT_ALIGN

    def export_full_report(self, filepath: str, data: Dict[str, Any],
                           company_name: str = "", fiscal_year: str = "",
                           ratios: Dict = None, tax_data: Dict = None,
                           budget_data: Dict = None, cashflow_data: Dict = None,
                           cost_centers: List = None,
                           comparative_data: Dict = None) -> bool:
        """تصدير تقرير كامل مع عدة أوراق"""
        try:
            from openpyxl import Workbook
            wb = Workbook()

            self._add_cover_sheet(wb, company_name, fiscal_year)
            self._add_balance_sheet(wb, data)
            self._add_income_statement(wb, data)
            if ratios:
                self._add_ratios_sheet(wb, ratios)
            if tax_data:
                self._add_tax_sheet(wb, tax_data, company_name)
            if cashflow_data:
                self._add_cashflow_sheet(wb, cashflow_data)
            if budget_data:
                self._add_budget_sheet(wb, budget_data)
            if cost_centers:
                self._add_cost_center_sheet(wb, cost_centers)
            if comparative_data:
                self._add_comparative_sheet(wb, comparative_data)

            self._remove_default_sheet(wb)
            wb.save(filepath)
            logger.info(f"Full Excel report saved to {filepath}")
            return True
        except Exception as e:
            logger.error(f"Full Excel export error: {e}")
            return False

    def _add_cover_sheet(self, wb, company_name: str, fiscal_year: str):
        ws = wb.active
        ws.title = "Cover"
        ws.sheet_properties.tabColor = "2980B9"

        ws.merge_cells('A1:F1')
        ws['A1'] = f"Financial Report - {company_name}" if company_name else "Financial Report"
        ws['A1'].font = self.TITLE_FONT
        ws['A1'].alignment = self.CENTER_ALIGN

        ws.merge_cells('A3:F3')
        ws['A3'] = f"Fiscal Year: {fiscal_year}" if fiscal_year else ""
        ws['A3'].font = self.SUBTITLE_FONT
        ws['A3'].alignment = self.CENTER_ALIGN

        ws.merge_cells('A4:F4')
        ws['A4'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws['A4'].font = self.SUBTITLE_FONT
        ws['A4'].alignment = self.CENTER_ALIGN

        ws.merge_cells('A6:F6')
        ws['A6'] = "Smart Accounting Platform - Algerian Financial Analysis"
        ws['A6'].font = self.Font(bold=True, size=12, color="2980B9")
        ws['A6'].alignment = self.CENTER_ALIGN

        for col in ['A', 'B', 'C', 'D', 'E', 'F']:
            ws.column_dimensions[col].width = 20

    def _add_balance_sheet(self, wb, data: Dict):
        ws = wb.create_sheet("Balance Sheet")
        ws.sheet_properties.tabColor = "27AE60"

        ws.merge_cells('A1:D1')
        ws['A1'] = "Balance Sheet / الميزانية العمومية"
        ws['A1'].font = self.TITLE_FONT

        headers = ["Account", "Current Year", "Previous Year", "Change %"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        self._style_header_row(ws, 3, 4)

        bs_items = [
            ("CURRENT ASSETS / الأصول المتداولة", None),
            ("  Cash and equivalents", data.get("cash", 0)),
            ("  Accounts receivable", data.get("accounts_receivable", 0)),
            ("  Inventory", data.get("inventory", 0)),
            ("  Total Current Assets", data.get("current_assets", 0)),
            ("", None),
            ("FIXED ASSETS / الأصول الثابتة", None),
            ("  Property, plant & equipment", data.get("fixed_assets", 0)),
            ("  Intangible assets", data.get("intangible_assets", 0)),
            ("  Total Fixed Assets", (data.get("fixed_assets", 0) + data.get("intangible_assets", 0))),
            ("", None),
            ("TOTAL ASSETS / إجمالي الأصول", data.get("total_assets", 0)),
            ("", None),
            ("CURRENT LIABILITIES / الخصوم المتداولة", None),
            ("  Accounts payable", data.get("accounts_payable", 0)),
            ("  Short-term debt", data.get("short_term_debt", 0)),
            ("  Total Current Liabilities", data.get("current_liabilities", 0)),
            ("", None),
            ("LONG-TERM LIABILITIES / الخصوم طويلة الأجل", None),
            ("  Long-term debt", data.get("long_term_debt", 0)),
            ("  Total Long-term Liabilities", data.get("long_term_liabilities", 0)),
            ("", None),
            ("EQUITY / حقوق الملكية", None),
            ("  Share capital", data.get("share_capital", 0)),
            ("  Retained earnings", data.get("retained_earnings", 0)),
            ("TOTAL EQUITY / إجمالي حقوق الملكية", data.get("equity", 0)),
            ("", None),
            ("TOTAL LIABILITIES & EQUITY", (data.get("total_liabilities", 0) + data.get("equity", 0))),
        ]

        row = 4
        for label, value in bs_items:
            ws.cell(row=row, column=1, value=label)
            if value is not None:
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=2).number_format = '#,##0.00'
            is_total = label.startswith("TOTAL") or label.startswith("  Total")
            self._style_data_row(ws, row, 4, is_total=is_total)
            row += 1

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 15

    def _add_income_statement(self, wb, data: Dict):
        ws = wb.create_sheet("Income Statement")
        ws.sheet_properties.tabColor = "E74C3C"

        ws.merge_cells('A1:D1')
        ws['A1'] = "Income Statement / قائمة الدخل"
        ws['A1'].font = self.TITLE_FONT

        headers = ["Account", "Current Year", "Previous Year", "Change %"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        self._style_header_row(ws, 3, 4)

        rev = data.get("revenue", 0)
        cogs = data.get("cost_of_goods_sold", 0)
        gross = data.get("gross_profit", rev - cogs)
        opex = data.get("operating_expenses", 0)
        ebit = gross - opex
        net = data.get("net_income", ebit)

        items = [
            ("REVENUE / الإيرادات", rev),
            ("COGS / تكلفة المبيعات", cogs),
            ("GROSS PROFIT / صافي الربح الإجمالي", gross),
            ("", None),
            ("OPERATING EXPENSES / المصاريف التشغيلية", None),
            ("  Salaries", data.get("salaries_expense", 0)),
            ("  Rent", data.get("rent_expense", 0)),
            ("  Utilities", data.get("utilities_expense", 0)),
            ("  Depreciation", data.get("depreciation", 0)),
            ("  Total Operating Expenses", opex),
            ("", None),
            ("EBIT / الربح قبل الفوائد والضرائب", ebit),
            ("Interest expense", data.get("interest_expense", 0)),
            ("EBT / الربح قبل الضرائب", ebit - data.get("interest_expense", 0)),
            ("Income tax", data.get("income_tax", 0)),
            ("NET INCOME / صافي الدخل", net),
        ]

        row = 4
        for label, value in items:
            ws.cell(row=row, column=1, value=label)
            if value is not None:
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=2).number_format = '#,##0.00'
            is_total = label.startswith("GROSS") or label.startswith("NET") or label.startswith("EBIT") or label.startswith("EBT") or label.startswith("  Total")
            self._style_data_row(ws, row, 4, is_total=is_total)
            row += 1

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20

        if rev > 0:
            try:
                from openpyxl.chart import PieChart, Reference
                chart = PieChart()
                chart.title = "Revenue vs COGS vs Expenses"
                chart_data = Reference(ws, min_col=2, min_row=4, max_row=6)
                chart_cats = Reference(ws, min_col=1, min_row=5, max_row=7)
                chart.add_data(chart_data, titles_from_data=False)
                chart.set_categories(chart_cats)
                chart.width = 16
                chart.height = 10
                ws.add_chart(chart, "D4")
            except Exception:
                logger.debug("Failed to add chart", exc_info=True)

    def _add_ratios_sheet(self, wb, ratios: Dict):
        ws = wb.create_sheet("Financial Ratios")
        ws.sheet_properties.tabColor = "8E44AD"

        ws.merge_cells('A1:C1')
        ws['A1'] = "Financial Ratios / النسب المالية"
        ws['A1'].font = self.TITLE_FONT

        categories = {
            "LIQUIDITY / السيولة": [
                ("current_ratio", "Current Ratio / نسبة التداول"),
                ("quick_ratio", "Quick Ratio / نسبة السرعة"),
                ("cash_ratio", "Cash Ratio /_ratio النقدية"),
            ],
            "PROFITABILITY / الربحية": [
                ("gross_profit_margin", "Gross Margin (%)"),
                ("net_profit_margin", "Net Margin (%)"),
                ("roa", "ROA (%)"),
                ("roe", "ROE (%)"),
            ],
            "LEVERAGE / الرفع المالي": [
                ("debt_to_equity", "Debt/Equity"),
                ("debt_ratio", "Debt Ratio"),
                ("interest_coverage", "Interest Coverage"),
            ],
            "EFFICIENCY / الكفاءة": [
                ("asset_turnover", "Asset Turnover"),
                ("inventory_turnover", "Inventory Turnover"),
                ("receivable_turnover", "Receivable Turnover"),
            ],
        }

        headers = ["Category", "Ratio", "Value"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        self._style_header_row(ws, 3, 3)

        row = 4
        for cat_name, items in categories.items():
            ws.cell(row=row, column=1, value=cat_name)
            ws.cell(row=row, column=1).font = self.Font(bold=True, size=11, color="2C3E50")
            row += 1
            for key, label in items:
                val = ratios.get(key, 0)
                ws.cell(row=row, column=1, value="")
                ws.cell(row=row, column=2, value=label)
                ws.cell(row=row, column=3, value=round(val, 4) if isinstance(val, float) else val)
                self._style_data_row(ws, row, 3)
                row += 1
            row += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 35
        ws.column_dimensions['C'].width = 15

    def _add_tax_sheet(self, wb, tax_data: Dict, company_name: str = ""):
        ws = wb.create_sheet("Tax Data")
        ws.sheet_properties.tabColor = "F39C12"

        ws.merge_cells('A1:C1')
        ws['A1'] = "Tax Summary / ملخص الضرائب"
        ws['A1'].font = self.TITLE_FONT

        headers = ["Tax Type / نوع الضريبة", "Amount (DZD)", "Rate"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        self._style_header_row(ws, 3, 3)

        taxes = [
            ("IBS / ضريبة الدخل للمؤسسات", tax_data.get("ibs", 0), "19%"),
            ("TVA / ضريبة القيمة المضافة", tax_data.get("tva", 0), "19%"),
            ("IRG / ضريبة الدخل الفردي", tax_data.get("irg", 0), "Progressive"),
            ("CNAS / الضمان الاجتماعي", tax_data.get("cnas", 0), "26%"),
            ("CNAC / التأمينات الاجتماعية", tax_data.get("cnac", 0), "2.25%"),
            ("Versement Forfaitaire", tax_data.get("versement_forfaitaire", 0), "2.5%"),
        ]

        row = 4
        for label, amount, rate in taxes:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=amount)
            ws.cell(row=row, column=3, value=rate)
            self._style_data_row(ws, row, 3)
            row += 1

        ws.cell(row=row, column=1, value="TOTAL / المجموع")
        ws.cell(row=row, column=2, value=tax_data.get("total_taxes", sum(t[1] for t in taxes)))
        self._style_data_row(ws, row, 3, is_total=True)

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

    def _add_cashflow_sheet(self, wb, cashflow: Dict):
        ws = wb.create_sheet("Cash Flow")
        ws.sheet_properties.tabColor = "1ABC9C"

        ws.merge_cells('A1:C1')
        ws['A1'] = "Cash Flow Statement / قائمة التدفقات النقدية"
        ws['A1'].font = self.TITLE_FONT

        headers = ["Item", "Amount (DZD)", "Category"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        self._style_header_row(ws, 3, 3)

        items = [
            ("Net income", cashflow.get("net_income", 0), "Operating"),
            ("Depreciation", cashflow.get("depreciation", 0), "Operating"),
            ("Changes in working capital", cashflow.get("working_capital", 0), "Operating"),
            ("Cash from operations", cashflow.get("operating", 0), "Operating"),
            ("", None, ""),
            ("Capital expenditures", cashflow.get("capex", 0), "Investing"),
            ("Asset sales", cashflow.get("asset_sales", 0), "Investing"),
            ("Cash from investing", cashflow.get("investing", 0), "Investing"),
            ("", None, ""),
            ("Debt issuance/repayment", cashflow.get("debt", 0), "Financing"),
            ("Dividends paid", cashflow.get("dividends", 0), "Financing"),
            ("Cash from financing", cashflow.get("financing", 0), "Financing"),
            ("", None, ""),
            ("Net change in cash", cashflow.get("net_change", 0), "Total"),
        ]

        row = 4
        for label, value, cat in items:
            ws.cell(row=row, column=1, value=label)
            if value is not None:
                ws.cell(row=row, column=2, value=value)
            ws.cell(row=row, column=3, value=cat)
            self._style_data_row(ws, row, 3, is_total=(cat == "Total"))
            row += 1

        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15

    def _add_budget_sheet(self, wb, budget: Dict):
        ws = wb.create_sheet("Budget vs Actual")
        ws.sheet_properties.tabColor = "3498DB"

        ws.merge_cells('A1:E1')
        ws['A1'] = "Budget vs Actual / الميزانية مقابل الفعلي"
        ws['A1'].font = self.TITLE_FONT

        headers = ["Account", "Budget", "Actual", "Variance", "Variance %"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        self._style_header_row(ws, 3, 5)

        items = budget.get("items", [
            {"name": "Revenue", "budget": 0, "actual": 0},
            {"name": "COGS", "budget": 0, "actual": 0},
            {"name": "Operating Expenses", "budget": 0, "actual": 0},
        ])

        row = 4
        for item in items:
            b = item.get("budget", 0)
            a = item.get("actual", 0)
            var = a - b
            var_pct = (var / b * 100) if b != 0 else 0

            ws.cell(row=row, column=1, value=item.get("name", ""))
            ws.cell(row=row, column=2, value=b)
            ws.cell(row=row, column=3, value=a)
            ws.cell(row=row, column=4, value=var)
            ws.cell(row=row, column=5, value=f"{var_pct:.1f}%")
            self._style_data_row(ws, row, 5)
            row += 1

        ws.column_dimensions['A'].width = 30
        for c in ['B', 'C', 'D', 'E']:
            ws.column_dimensions[c].width = 18

    def _add_cost_center_sheet(self, wb, cost_centers: List):
        ws = wb.create_sheet("Cost Centers")
        ws.sheet_properties.tabColor = "E67E22"

        ws.merge_cells('A1:D1')
        ws['A1'] = "Cost Centers / مراكز التكلفة"
        ws['A1'].font = self.TITLE_FONT

        headers = ["Center Name", "Budget", "Actual", "Efficiency %"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        self._style_header_row(ws, 3, 4)

        row = 4
        for cc in cost_centers:
            ws.cell(row=row, column=1, value=cc.get("name", ""))
            ws.cell(row=row, column=2, value=cc.get("budget", 0))
            ws.cell(row=row, column=3, value=cc.get("actual", 0))
            eff = cc.get("efficiency", 0)
            ws.cell(row=row, column=4, value=f"{eff:.1f}%")
            self._style_data_row(ws, row, 4)
            row += 1

        ws.column_dimensions['A'].width = 30
        for c in ['B', 'C', 'D']:
            ws.column_dimensions[c].width = 18

    def _add_comparative_sheet(self, wb, comp: Dict):
        ws = wb.create_sheet("Comparative Analysis")
        ws.sheet_properties.tabColor = "9B59B6"

        ws.merge_cells('A1:D1')
        ws['A1'] = "Comparative Analysis / التحليل المقارن"
        ws['A1'].font = self.TITLE_FONT

        years = comp.get("years", ["Year 1", "Year 2"])
        headers = ["Metric"] + years
        for col, h in enumerate(headers, 1):
            ws.cell(row=3, column=col, value=h)
        self._style_header_row(ws, 3, len(headers))

        metrics = comp.get("metrics", {})
        row = 4
        for metric_name, values in metrics.items():
            ws.cell(row=row, column=1, value=metric_name)
            for i, val in enumerate(values):
                ws.cell(row=row, column=2 + i, value=val)
            self._style_data_row(ws, row, len(headers))
            row += 1

        ws.column_dimensions['A'].width = 30
        for i in range(len(years)):
            ws.column_dimensions[chr(66 + i)].width = 18

    def _remove_default_sheet(self, wb):
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    def export_comparison(self, data: List[Dict], filepath: str,
                          title: str = "Benchmark Comparison") -> bool:
        """تصدير تقرير المقارنة الصناعية"""
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Benchmark"
            ws.sheet_properties.tabColor = "2980B9"

            ws.merge_cells('A1:G1')
            ws['A1'] = title
            ws['A1'].font = self.TITLE_FONT
            ws['A1'].alignment = self.CENTER_ALIGN

            ws.merge_cells('A2:G2')
            ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
            ws['A2'].font = self.SUBTITLE_FONT
            ws['A2'].alignment = self.CENTER_ALIGN

            headers = ["Ratio", "Company Value", "Sector Min",
                       "Sector Avg", "Sector Max", "Status", "Score"]
            for col, h in enumerate(headers, 1):
                ws.cell(row=4, column=col, value=h)
            self._style_header_row(ws, 4, 7)

            row = 5
            for item in data:
                ws.cell(row=row, column=1, value=item.get("Ratio", ""))
                for col_idx, key in enumerate(
                    ["Company Value", "Sector Min", "Sector Avg",
                     "Sector Max", "Status", "Score"], 2
                ):
                    ws.cell(row=row, column=col_idx, value=item.get(key, ""))
                self._style_data_row(ws, row, 7)
                row += 1

            ws.column_dimensions['A'].width = 25
            for c in ['B', 'C', 'D', 'E', 'F', 'G']:
                ws.column_dimensions[c].width = 16

            self._remove_default_sheet(wb)
            wb.save(filepath)
            logger.info(f"Benchmark comparison saved to {filepath}")
            return True
        except Exception as e:
            logger.error(f"Export comparison error: {e}")
            return False


excel_exporter = ExcelExporter()
