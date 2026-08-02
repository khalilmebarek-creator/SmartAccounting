# محرك اختبار المستخدمين الحقيقيين وجمع الملاحظات
# =================================================
# - مجموعات المستخدمين: محاسبون / مديرون / ملاك أعمال / مديرون ماليون (CFO)
# - سيناريوهات الاختبار: إدخال بيانات / توليد تقارير / تحليل / اتخاذ قرار / حل مشكلات
# - جمع الملاحظات: سهولة استخدام / أداء / ميزات / أخطاء / اقتراحات
# - المخرجات: تقرير الملاحظات + قائمة المشكلات + طلبات التحسين + درجة رضا المستخدم
# التخزين: ذاكرة + تصدير/استيراد JSON + حفظ اختياري في SQLite

import csv
import json
import os
import tempfile
import time
import uuid

from database.db_connection import get_connection
from utils.app_logger import get_logger

log = get_logger("user_testing")

USER_GROUPS = ("accountant", "manager", "business_owner", "cfo")
SCENARIOS = ("data_entry", "report_generation", "analysis", "decision_making", "issue_resolution")
FEEDBACK_CATEGORIES = ("usability", "performance", "features", "bugs", "suggestions")
PRIORITIES = ("low", "medium", "high", "critical")
STATUSES = ("open", "in_progress", "resolved", "closed")

PRIORITY_WEIGHT = {"critical": 4, "high": 3, "medium": 2, "low": 1}

# درجة الرضا: عتبات تصنيف (ممتاز ≥4.5 / جيد ≥3.5 / متوسط ≥2.5 / ضعيف <2.5)
SATISFACTION_LEVELS = (
    ("excellent", 4.5),
    ("good", 3.5),
    ("average", 2.5),
    ("poor", 0.0),
)

SESSION_TABLE = "user_testing_sessions"
APP_ID = "SmartAccounting"
FORMAT_VERSION = 1

DEMO_TESTS = {
    "accountant": "محاسب أول",
    "manager": "مدير مالي",
    "business_owner": "مالك الشركة",
    "cfo": "مدير مالي تنفيذي (CFO)",
}


def _now():
    return time.strftime("%Y-%m-%d %H:%M:%S")


def _atomic_write(path, text):
    dir_name = os.path.dirname(path) or "."
    fd, tmp_path = tempfile.mkstemp(dir=dir_name, suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            f.write(text)
        os.replace(tmp_path, path)
    except Exception:
        try:
            os.remove(tmp_path)
        except OSError:
            pass
        raise


def _safe_json_read(path, default=None):
    if not os.path.exists(path):
        return default
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception as e:
        log.error("Failed to read %s: %s", path, e)
        return default


def satisfaction_level(score):
    """تصنيف درجة الرضا (0-5) إلى مستوى نصي."""
    level = "poor"
    for name, threshold in reversed(SATISFACTION_LEVELS):
        if score >= threshold:
            level = name
    return level


# أحرف لا يوفّرها خط Amiri في الـ PDF → استبدالها بمكافئات آمنة
_PDF_SAFE_CHARS = str.maketrans({
    "\u2550": "=", "\u2500": "-", "\u2551": "|", "\u2502": "|",
    "\u251c": "+", "\u2514": "+",
    "\u2014": "-", "\u2013": "-", "\u2022": "-",
    "\u201c": '"', "\u201d": '"', "\u00ab": '"', "\u00bb": '"',
    "\u2705": "OK", "\u274c": "X", "\u26a0": "!", "\u2139": "i",
})


def _sanitize_for_pdf(text):
    return text.translate(_PDF_SAFE_CHARS)


def _wrap_line(pdf, text, max_width):
    """لف سطر نصي يدوياً حسب عرض المتاح (لأن multi_cell يفشل مع النصوص العربية)."""
    if not text:
        return [""]
    if pdf.get_string_width(text) <= max_width:
        return [text]
    words = text.split()
    lines, current = [], ""
    for word in words:
        test = (current + " " + word).strip()
        if pdf.get_string_width(test) <= max_width:
            current = test
        else:
            if current:
                lines.append(current)
            current = word
    if current:
        lines.append(current)
    return lines


class UserTestingEngine:
    """محرك اختبار المستخدمين الحقيقيين — جلسات + ملاحظات + درجات + تقارير."""

    def __init__(self):
        self._sessions = {}
        self._next_id = 1

    # ==================== الجلسات ====================

    def create_session(self, name, tester_name="", user_group="accountant",
                       scenario="data_entry", environment="", notes=""):
        """إنشاء جلسة اختبار جديدة وإرجاعها."""
        if not (name and str(name).strip()):
            raise ValueError("session name required")
        self._validate_member(user_group, USER_GROUPS, "user_group")
        self._validate_member(scenario, SCENARIOS, "scenario")
        session_id = str(self._next_id)
        self._next_id += 1
        session = {
            "id": session_id,
            "name": str(name).strip(),
            "tester_name": tester_name or "",
            "user_group": user_group,
            "scenario": scenario,
            "environment": environment or "",
            "notes": notes or "",
            "created_at": _now(),
            "feedback": [],
        }
        self._sessions[session_id] = session
        log.info("Created testing session %s (%s)", session_id, name)
        return session

    def get_session(self, session_id):
        return self._sessions.get(str(session_id))

    def list_sessions(self):
        return list(self._sessions.values())

    def delete_session(self, session_id):
        return self._sessions.pop(str(session_id), None) is not None

    def _validate_member(self, value, allowed, field):
        if value not in allowed:
            raise ValueError(f"invalid {field}: {value!r}")

    # ==================== الملاحظات ====================

    def add_feedback(self, session_id, category, comment, rating,
                     user_group=None, scenario=None, priority="medium",
                     title="", status="open"):
        """إضافة ملاحظة إلى جلسة اختبار مع التحقق من الصحة."""
        session = self.get_session(session_id)
        if not session:
            raise KeyError(f"session not found: {session_id}")
        self._validate_member(category, FEEDBACK_CATEGORIES, "category")
        self._validate_member(priority, PRIORITIES, "priority")
        self._validate_member(status, STATUSES, "status")
        rating = self._validate_rating(rating)
        comment = (comment or "").strip()
        if not comment:
            raise ValueError("comment required")
        group = user_group or session["user_group"]
        scn = scenario or session["scenario"]
        self._validate_member(group, USER_GROUPS, "user_group")
        self._validate_member(scn, SCENARIOS, "scenario")
        feedback = {
            "id": str(uuid.uuid4()),
            "user_group": group,
            "scenario": scn,
            "category": category,
            "priority": priority,
            "status": status,
            "rating": rating,
            "title": (title or "").strip(),
            "comment": comment,
            "created_at": _now(),
        }
        session["feedback"].append(feedback)
        log.info("Added %s feedback to session %s (rating=%s)", category, session_id, rating)
        return feedback

    def update_feedback(self, session_id, feedback_id, **fields):
        """تحديث حقول ملاحظة (مثل الحالة/الأولوية/التقييم)."""
        feedback = self._find_feedback(session_id, feedback_id)
        allowed = {"category", "priority", "status", "rating", "title",
                   "comment", "user_group", "scenario"}
        for key, value in fields.items():
            if key not in allowed:
                raise KeyError(f"unknown field: {key}")
            if key == "rating":
                value = self._validate_rating(value)
            elif key == "priority":
                self._validate_member(value, PRIORITIES, "priority")
            elif key == "status":
                self._validate_member(value, STATUSES, "status")
            elif key == "category":
                self._validate_member(value, FEEDBACK_CATEGORIES, "category")
            elif key == "user_group":
                self._validate_member(value, USER_GROUPS, "user_group")
            elif key == "scenario":
                self._validate_member(value, SCENARIOS, "scenario")
            feedback[key] = value
        return feedback

    def delete_feedback(self, session_id, feedback_id):
        session = self.get_session(session_id)
        if not session:
            raise KeyError(f"session not found: {session_id}")
        for i, feedback in enumerate(session["feedback"]):
            if feedback["id"] == feedback_id:
                session["feedback"].pop(i)
                return True
        return False

    def _find_feedback(self, session_id, feedback_id):
        session = self.get_session(session_id)
        if not session:
            raise KeyError(f"session not found: {session_id}")
        for feedback in session["feedback"]:
            if feedback["id"] == feedback_id:
                return feedback
        raise KeyError(f"feedback not found: {feedback_id}")

    def list_feedback(self, session_id, category=None, user_group=None,
                      status=None):
        """قائمة الملاحظات مع تصفية اختيارية."""
        session = self.get_session(session_id)
        if not session:
            return []
        items = session["feedback"]
        if category:
            items = [f for f in items if f["category"] == category]
        if user_group:
            items = [f for f in items if f["user_group"] == user_group]
        if status:
            items = [f for f in items if f["status"] == status]
        return items

    @staticmethod
    def _validate_rating(rating):
        try:
            rating = int(rating)
        except (TypeError, ValueError):
            raise ValueError(f"invalid rating: {rating!r}") from None
        if not 1 <= rating <= 5:
            raise ValueError("rating must be between 1 and 5")
        return rating

    # ==================== درجة رضا المستخدم ====================

    def satisfaction_score(self, session_id):
        """درجة الرضا الإجمالية + التفصيل حسب التصنيف/المجموعة/السيناريو."""
        session = self.get_session(session_id)
        if not session:
            return {"overall": 0.0, "count": 0, "level": "poor",
                    "by_category": {}, "by_user_group": {}, "by_scenario": {}}
        items = session["feedback"]
        if not items:
            return {"overall": 0.0, "count": 0, "level": "poor",
                    "by_category": {}, "by_user_group": {}, "by_scenario": {}}
        overall = sum(f["rating"] for f in items) / len(items)
        return {
            "overall": round(overall, 2),
            "count": len(items),
            "level": satisfaction_level(overall),
            "by_category": self._breakdown(items, "category"),
            "by_user_group": self._breakdown(items, "user_group"),
            "by_scenario": self._breakdown(items, "scenario"),
        }

    @staticmethod
    def _breakdown(items, key):
        groups = {}
        for item in items:
            groups.setdefault(item[key], []).append(item["rating"])
        return {k: round(sum(v) / len(v), 2) for k, v in sorted(groups.items())}

    # ==================== التقارير ====================

    def feedback_report(self, session_id):
        """تقرير الملاحظات: إحصائيات + توزيع + متوسطات + ملاحظات كاملة."""
        session = self.get_session(session_id)
        if not session:
            return {}
        items = session["feedback"]
        score = self.satisfaction_score(session_id)
        counts_by_category = {c: 0 for c in FEEDBACK_CATEGORIES}
        counts_by_status = {s: 0 for s in STATUSES}
        counts_by_priority = {p: 0 for p in PRIORITIES}
        for item in items:
            counts_by_category[item["category"]] += 1
            counts_by_status[item["status"]] += 1
            counts_by_priority[item["priority"]] += 1
        return {
            "session": {
                "id": session["id"], "name": session["name"],
                "tester_name": session["tester_name"],
                "user_group": session["user_group"],
                "scenario": session["scenario"],
                "environment": session["environment"],
                "notes": session["notes"],
                "created_at": session["created_at"],
            },
            "total_feedback": len(items),
            "counts_by_category": counts_by_category,
            "counts_by_status": counts_by_status,
            "counts_by_priority": counts_by_priority,
            "avg_by_category": score["by_category"],
            "open_issues": sum(1 for f in items
                               if f["category"] == "bugs" and f["status"] != "resolved"),
            "enhancement_requests": sum(1 for f in items
                                        if f["category"] in ("features", "suggestions")),
            "satisfaction": score,
            "feedback": list(items),
        }

    def issue_list(self, session_id):
        """قائمة المشكلات: ملاحظات 'أخطاء' مرتبة حسب الأولوية."""
        session = self.get_session(session_id)
        if not session:
            return []
        issues = [f for f in session["feedback"] if f["category"] == "bugs"]
        issues.sort(key=lambda f: (-PRIORITY_WEIGHT[f["priority"]],
                                   f["created_at"]))
        return issues

    def enhancement_requests(self, session_id):
        """طلبات التحسين: ملاحظات 'ميزات' و'اقتراحات' مرتبة حسب الأولوية."""
        session = self.get_session(session_id)
        if not session:
            return []
        items = [f for f in session["feedback"]
                 if f["category"] in ("features", "suggestions")]
        items.sort(key=lambda f: (-PRIORITY_WEIGHT[f["priority"]],
                                  f["created_at"]))
        return items

    def summary_text(self, session_id):
        """نص تقرير ملخص (للإشعار/الـ PDF/التصدير النصي)."""
        session = self.get_session(session_id)
        if not session:
            return "لا توجد جلسة اختبار."
        report = self.feedback_report(session_id)
        score = report["satisfaction"]
        lines = [
            "═" * 60,
            f"تقرير اختبار المستخدمين: {session['name']}",
            f"المختبر: {session['tester_name'] or '—'}  |  المجموعة: {session['user_group']}  |  السيناريو: {session['scenario']}",
            f"التاريخ: {session['created_at']}  |  البيئة: {session['environment'] or '—'}",
            "═" * 60,
            f"إجمالي الملاحظات: {report['total_feedback']}",
            f"درجة الرضا: {score['overall']}/5  ({score['level']})",
            f"مشكلات مفتوحة: {report['open_issues']}  |  طلبات تحسين: {report['enhancement_requests']}",
            "",
            "التوزيع حسب التصنيف:",
        ]
        for category, count in report["counts_by_category"].items():
            lines.append(f"  • {category}: {count}")
        if report["feedback"]:
            lines += ["", "الملاحظات:"]
            for item in report["feedback"]:
                lines.append(
                    f"  [{item['category']}/{item['priority']}/{item['rating']}/5] "
                    f"{item['title'] or ''} — {item['comment']} ({item['status']})"
                )
        lines.append("")
        lines.append("═" * 60)
        return "\n".join(lines)

    # ==================== تصدير/استيراد JSON ====================

    def export_json(self, path):
        """تصدير كل الجلسات إلى ملف JSON واحد."""
        payload = {
            "app": APP_ID,
            "format": FORMAT_VERSION,
            "exported_at": _now(),
            "sessions": [self.to_dict(s) for s in self._sessions.values()],
        }
        _atomic_write(path, json.dumps(payload, ensure_ascii=False, indent=2))
        log.info("Exported %d testing sessions to %s", len(payload["sessions"]), path)
        return True

    def import_json(self, path):
        """استيراد الجلسات من ملف JSON وإرجاع عددها."""
        data = _safe_json_read(path)
        if not isinstance(data, dict) or "sessions" not in data:
            return 0
        count = 0
        for session in data["sessions"]:
            if self._import_session(session):
                count += 1
        log.info("Imported %d testing sessions from %s", count, path)
        return count

    def _import_session(self, session):
        if not isinstance(session, dict) or not session.get("name"):
            return False
        session_id = str(session.get("id") or self._next_id)
        if session_id in self._sessions:
            session_id = str(self._next_id)
        try:
            feedback = [
                {**f, "rating": self._validate_rating(f.get("rating", 0)),
                 "category": f.get("category"), "priority": f.get("priority", "medium"),
                 "status": f.get("status", "open")}
                for f in session.get("feedback") or []
            ]
        except ValueError:
            return False
        self._sessions[session_id] = {
            "id": session_id,
            "name": str(session["name"]).strip(),
            "tester_name": session.get("tester_name", ""),
            "user_group": session.get("user_group", "accountant"),
            "scenario": session.get("scenario", "data_entry"),
            "environment": session.get("environment", ""),
            "notes": session.get("notes", ""),
            "created_at": session.get("created_at", _now()),
            "feedback": feedback,
        }
        try:
            self._next_id = max(self._next_id, int(session_id) + 1)
        except (TypeError, ValueError):
            pass
        return True

    @staticmethod
    def to_dict(session):
        return {
            "id": session["id"],
            "name": session["name"],
            "tester_name": session["tester_name"],
            "user_group": session["user_group"],
            "scenario": session["scenario"],
            "environment": session["environment"],
            "notes": session["notes"],
            "created_at": session["created_at"],
            "feedback": list(session["feedback"]),
        }

    # ==================== الحفظ في قاعدة البيانات ====================

    def _ensure_table(self, conn):
        conn.execute(f"""
            CREATE TABLE IF NOT EXISTS {SESSION_TABLE} (
                session_id TEXT PRIMARY KEY,
                name TEXT NOT NULL,
                payload TEXT NOT NULL,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        """)

    def save_session_db(self, session_id):
        """حفظ جلسة كاملة في قاعدة SQLite (payload JSON)."""
        session = self.get_session(session_id)
        if not session:
            return False
        payload = json.dumps(self.to_dict(session), ensure_ascii=False)
        try:
            with get_connection() as conn:
                self._ensure_table(conn)
                conn.execute(
                    f"INSERT INTO {SESSION_TABLE} (session_id, name, payload) VALUES (?, ?, ?) "
                    f"ON CONFLICT(session_id) DO UPDATE SET name=excluded.name, "
                    f"payload=excluded.payload, updated_at=CURRENT_TIMESTAMP",
                    (session["id"], session["name"], payload),
                )
            log.info("Saved session %s to database", session_id)
            return True
        except Exception as e:
            log.error("save_session_db error: %s", e)
            return False

    def load_session_db(self, session_id):
        """تحميل جلسة من قاعدة SQLite إلى الذاكرة."""
        row = None
        try:
            with get_connection() as conn:
                if not conn.table_exists(SESSION_TABLE):
                    return False
                row = conn.fetch_one(
                    f"SELECT payload FROM {SESSION_TABLE} WHERE session_id = ?",
                    (str(session_id),),
                )
        except Exception as e:
            log.error("load_session_db error: %s", e)
            return False
        if not row:
            return False
        data = json.loads(row[0])
        return self._import_session(data)

    def list_session_ids_db(self):
        try:
            with get_connection() as conn:
                if not conn.table_exists(SESSION_TABLE):
                    return []
                rows = conn.fetch_all(
                    f"SELECT session_id, name FROM {SESSION_TABLE} ORDER BY name"
                )
        except Exception as e:
            log.error("list_session_ids_db error: %s", e)
            return []
        return [{"id": r[0], "name": r[1]} for r in rows]

    def delete_session_db(self, session_id):
        try:
            with get_connection() as conn:
                if not conn.table_exists(SESSION_TABLE):
                    return False
                conn.execute(
                    f"DELETE FROM {SESSION_TABLE} WHERE session_id = ?",
                    (str(session_id),),
                )
            return True
        except Exception as e:
            log.error("delete_session_db error: %s", e)
            return False

    # ==================== تصدير Excel / PDF / CSV ====================

    def export_excel(self, path, session_id=None):
        """تصدير الملاحظات والتلخيص إلى Excel (أوراق متعددة)."""
        sessions = self._target_sessions(session_id)
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            log.warning("openpyxl not available")
            return False
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Feedback"
            header = ["Session", "Date", "Group", "Scenario", "Category",
                      "Priority", "Status", "Rating", "Title", "Comment"]
            ws.append(header)
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
            for session in sessions:
                for f in session["feedback"]:
                    ws.append([session["name"], f["created_at"], f["user_group"],
                               f["scenario"], f["category"], f["priority"],
                               f["status"], f["rating"], f["title"], f["comment"]])
            summary = wb.create_sheet("Summary")
            summary.append(["Session", "Feedback Count", "Satisfaction", "Level"])
            for cell in summary[1]:
                cell.font = Font(bold=True)
            for session in sessions:
                score = self.satisfaction_score(session["id"])
                summary.append([session["name"], score["count"],
                                score["overall"], score["level"]])
            wb.save(path)
            log.info("Excel report saved to %s", path)
            return True
        except Exception as e:
            log.error("Excel export error: %s", e)
            return False

    def export_csv(self, path, session_id=None):
        """تصدير الملاحظات إلى CSV."""
        sessions = self._target_sessions(session_id)
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["Session", "Date", "Group", "Scenario",
                                 "Category", "Priority", "Status", "Rating",
                                 "Title", "Comment"])
                for session in sessions:
                    for item in session["feedback"]:
                        writer.writerow([session["name"], item["created_at"],
                                         item["user_group"], item["scenario"],
                                         item["category"], item["priority"],
                                         item["status"], item["rating"],
                                         item["title"], item["comment"]])
            log.info("CSV exported to %s", path)
            return True
        except Exception as e:
            log.error("CSV export error: %s", e)
            return False

    def export_pdf(self, path, session_id=None):
        """تصدير تقرير ملخص إلى PDF مع دعم الخط العربي."""
        sessions = self._target_sessions(session_id)
        try:
            from fpdf import FPDF
        except ImportError:
            log.warning("fpdf not available")
            return False
        font_path = self._arabic_font_path()
        try:
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=20)
            if font_path:
                pdf.add_font("Amiri", "", font_path, uni=True)
            pdf.add_page()
            max_width = pdf.epw
            for session in sessions:
                text = self.summary_text(session["id"])
                if font_path:
                    pdf.set_font("Amiri", size=11)
                else:
                    pdf.set_font("Helvetica", size=11)
                for line in _sanitize_for_pdf(text).splitlines():
                    if not font_path:
                        line = line.encode("cp1252", errors="replace").decode("cp1252")
                    for wrapped in _wrap_line(pdf, line, max_width):
                        pdf.cell(0, 7, wrapped, ln=True)
                pdf.ln(4)
            pdf.output(path)
            log.info("PDF report saved to %s", path)
            return True
        except Exception as e:
            log.error("PDF export error: %s", e)
            return False

    def _target_sessions(self, session_id):
        if session_id is not None:
            session = self.get_session(session_id)
            return [session] if session else []
        return list(self._sessions.values())

    @staticmethod
    def _arabic_font_path():
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        font_path = os.path.join(base, "ui", "resources", "fonts", "Amiri-Regular.ttf")
        return font_path if os.path.exists(font_path) else None

    # ==================== بيانات تجريبية ====================

    def build_demo_data(self):
        """إنشاء جلسات تجريبية لكل مجموعة مستخدمين عبر كل السيناريوهات."""
        if self._sessions:
            return len(self._sessions)
        demo_feedback = {
            "accountant": [
                ("usability", 4, "medium", "إدخال البيانات سريع بعد التعوّد على الحقول", "data_entry"),
                ("bugs", 3, "high", "حقل تكلفة المبيعات يُعاد ضبطه عند تبديل السنة", "data_entry"),
                ("performance", 4, "low", "حساب النسب فوري حتى مع أرقام كبيرة", "report_generation"),
            ],
            "manager": [
                ("features", 5, "medium", "أتمنى تقارير أسبوعية تلقائية بالبريد", "report_generation"),
                ("usability", 4, "low", "الرسوم البيانية واضحة وسهلة القراءة", "analysis"),
                ("suggestions", 5, "medium", "إضافة مقارنة مع منافسين محليين في لوحة التحكم", "decision_making"),
            ],
            "business_owner": [
                ("usability", 5, "low", "واجهة بسيطة ولا أحتاج خبيراً محاسبياً", "analysis"),
                ("bugs", 2, "critical", "تعارض عند فتح شاشتين معاً يفقد التعديلات", "issue_resolution"),
                ("performance", 3, "medium", "بطء بسيط عند توليد تقرير 12 شهراً", "report_generation"),
            ],
            "cfo": [
                ("features", 5, "high", "نحتاج تنبؤات سيولة يومية وإشعارات", "decision_making"),
                ("suggestions", 4, "medium", "تصدير Excel يتضمن خيار سيناريوهات متعددة", "report_generation"),
                ("performance", 5, "low", "أداء ممتاز حتى مع ملف قاعدة كبير", "analysis"),
            ],
        }
        scenarios = SCENARIOS
        for group in USER_GROUPS:
            name = f"اختبار {DEMO_TESTS[group]} — {group}"
            session = self.create_session(
                name=name,
                tester_name=DEMO_TESTS[group],
                user_group=group,
                scenario=scenarios[list(USER_GROUPS).index(group) % len(scenarios)],
                environment="Windows 11 / v3.1",
                notes="جلسة تجريبية لتجربة محرك اختبار المستخدمين",
            )
            for category, rating, priority, comment, scenario in demo_feedback[group]:
                self.add_feedback(
                    session["id"], category=category, comment=comment,
                    rating=rating, priority=priority, scenario=scenario,
                    title="", status="open",
                )
        log.info("Built demo testing data (%d sessions)", len(self._sessions))
        return len(self._sessions)


user_testing_engine = UserTestingEngine()
