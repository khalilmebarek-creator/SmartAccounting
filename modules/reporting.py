# توليد التقارير
# ===============

import os
import logging
from datetime import datetime

log = logging.getLogger("reporting")

class ReportGenerator:
    """فئة لتوليد التقارير المالية"""
    
    def __init__(self, company_name, fiscal_year):
        """تهيئة مولد التقارير"""
        self.company_name = company_name
        self.fiscal_year = fiscal_year
        self.generated_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    def generate_balance_sheet_report(self, assets, liabilities, equity):
        """توليد تقرير الميزانية"""
        
        report = "\n" + "="*70
        report += "\n📊 قائمة المركز المالي (الميزانية)"
        report += f"\nالشركة: {self.company_name}"
        report += f"\nالسنة المالية: {self.fiscal_year}"
        report += f"\nتاريخ التقرير: {self.generated_date}"
        report += "\n" + "="*70
        
        report += "\n\n🏢 الأصول:"
        report += "\n" + "-"*70
        report += f"\n{'البيان':<40} {'المبلغ':<20}"
        report += "\n" + "-"*70
        
        if isinstance(assets, dict):
            total = 0
            for asset_name, asset_value in assets.items():
                report += f"\n{asset_name:<40} {asset_value:>15,.2f}"
                total += asset_value
            report += "\n" + "-"*70
            report += f"\n{'إجمالي الأصول':<40} {total:>15,.2f}"
        else:
            report += f"\n{'إجمالي الأصول':<40} {assets:>15,.2f}"
        
        report += "\n\n💳 الالتزامات:"
        report += "\n" + "-"*70
        report += f"\n{'البيان':<40} {'المبلغ':<20}"
        report += "\n" + "-"*70
        
        if isinstance(liabilities, dict):
            total_liab = 0
            for liab_name, liab_value in liabilities.items():
                report += f"\n{liab_name:<40} {liab_value:>15,.2f}"
                total_liab += liab_value
            report += "\n" + "-"*70
            report += f"\n{'إجمالي الالتزامات':<40} {total_liab:>15,.2f}"
        else:
            report += f"\n{'إجمالي الالتزامات':<40} {liabilities:>15,.2f}"
            total_liab = liabilities
        
        report += "\n\n💰 حقوق المالكين:"
        report += "\n" + "-"*70
        report += f"\n{'البيان':<40} {'المبلغ':<20}"
        report += "\n" + "-"*70
        
        if isinstance(equity, dict):
            total_eq = 0
            for eq_name, eq_value in equity.items():
                report += f"\n{eq_name:<40} {eq_value:>15,.2f}"
                total_eq += eq_value
            report += "\n" + "-"*70
            report += f"\n{'إجمالي حقوق المالكين':<40} {total_eq:>15,.2f}"
        else:
            report += f"\n{'إجمالي حقوق المالكين':<40} {equity:>15,.2f}"
            total_eq = equity
        
        report += "\n" + "="*70 + "\n"
        
        return report
    
    def generate_income_statement_report(self, revenue, cogs, expenses, net_income):
        """توليد تقرير الدخل"""
        
        report = "\n" + "="*70
        report += "\n💹 قائمة الدخل"
        report += f"\nالشركة: {self.company_name}"
        report += f"\nالسنة المالية: {self.fiscal_year}"
        report += f"\nتاريخ التقرير: {self.generated_date}"
        report += "\n" + "="*70
        
        report += "\n\n📈 البيان المالي:"
        report += "\n" + "-"*70
        report += f"\n{'البيان':<40} {'المبلغ':<20}"
        report += "\n" + "-"*70
        
        report += f"\n{'الإيرادات':<40} {revenue:>15,.2f}"
        report += f"\n{'تكلفة البضاعة المباعة':<40} ({cogs:>14,.2f})"
        
        gross_profit = revenue - cogs
        report += f"\n{'الربح الإجمالي':<40} {gross_profit:>15,.2f}"
        
        report += f"\n{'المصاريف التشغيلية':<40} ({expenses:>14,.2f})"
        
        operating_income = gross_profit - expenses
        report += f"\n{'الربح التشغيلي':<40} {operating_income:>15,.2f}"
        
        report += "\n" + "-"*70
        report += f"\n{'صافي الربح':<40} {net_income:>15,.2f}"
        report += "\n" + "="*70 + "\n"
        
        return report
    
    def generate_financial_ratios_report(self, ratios):
        """توليد تقرير النسب المالية"""
        
        report = "\n" + "="*70
        report += "\n📊 تقرير النسب المالية"
        report += f"\nالشركة: {self.company_name}"
        report += f"\nالسنة المالية: {self.fiscal_year}"
        report += f"\nتاريخ التقرير: {self.generated_date}"
        report += "\n" + "="*70
        
        report += "\n\n📈 نسب السيولة:"
        report += "\n" + "-"*70
        if 'current_ratio' in ratios:
            report += f"\nنسبة السيولة الحالية (Current Ratio): {ratios['current_ratio']}"
        if 'quick_ratio' in ratios:
            report += f"\nالنسبة السريعة (Quick Ratio): {ratios['quick_ratio']}"
        
        report += "\n\n💰 نسب الربحية:"
        report += "\n" + "-"*70
        if 'gross_profit_margin' in ratios:
            report += f"\nهامش الربح الإجمالي: {ratios['gross_profit_margin']}%"
        if 'net_profit_margin' in ratios:
            report += f"\nهامش صافي الربح: {ratios['net_profit_margin']}%"
        if 'roa' in ratios:
            report += f"\nالعائد على الأصول (ROA): {ratios['roa']}%"
        if 'roe' in ratios:
            report += f"\nالعائد على حقوق المالكين (ROE): {ratios['roe']}%"
        
        report += "\n\n⚙️  نسب الكفاءة:"
        report += "\n" + "-"*70
        if 'asset_turnover' in ratios:
            report += f"\nمعدل دوران الأصول: {ratios['asset_turnover']}"
        if 'receivables_turnover' in ratios:
            report += f"\nمعدل دوران الذمم المدينة: {ratios['receivables_turnover']}"
        if 'inventory_turnover' in ratios:
            report += f"\nمعدل دوران المخزون: {ratios['inventory_turnover']}"
        
        report += "\n\n📊 نسب الاستدانة:"
        report += "\n" + "-"*70
        if 'debt_to_equity' in ratios:
            report += f"\nنسبة الدين إلى حقوق المالكين: {ratios['debt_to_equity']}"
        if 'debt_ratio' in ratios:
            report += f"\nنسبة الدين: {ratios['debt_ratio']}"
        
        report += "\n" + "="*70 + "\n"
        
        return report
    
    def generate_comprehensive_report(self, balance_sheet, income_statement, ratios, analysis):
        """توليد تقرير شامل يجمع كل شيء"""
        
        report = "\n" + "█"*70
        report += "\n█" + " "*68 + "█"
        report += "\n█  التقرير المالي الشامل".center(68) + "█"
        report += "\n█" + " "*68 + "█"
        report += "\n█"*70
        
        report += f"\n\nالشركة: {self.company_name}"
        report += f"\nالسنة المالية: {self.fiscal_year}"
        report += f"\nتاريخ التقرير: {self.generated_date}"
        
        report += "\n\n" + "="*70
        report += "\n1️⃣ الميزانية"
        report += "\n" + "="*70
        if isinstance(balance_sheet, dict):
            for key, value in balance_sheet.items():
                if isinstance(value, (int, float)):
                    report += f"\n{key}: {value:,.2f}"
        
        report += "\n\n" + "="*70
        report += "\n2️⃣ قائمة الدخل"
        report += "\n" + "="*70
        if isinstance(income_statement, dict):
            for key, value in income_statement.items():
                if isinstance(value, (int, float)):
                    report += f"\n{key}: {value:,.2f}"
        
        report += "\n\n" + "="*70
        report += "\n3️⃣ النسب المالية"
        report += "\n" + "="*70
        if isinstance(ratios, dict):
            for key, value in ratios.items():
                if isinstance(value, (int, float)):
                    report += f"\n{key}: {value}"
        
        report += "\n\n" + "="*70
        report += "\n4️⃣ التحليل والملاحظات"
        report += "\n" + "="*70
        if isinstance(analysis, str):
            report += f"\n{analysis}"
        
        report += "\n\n" + "█"*70
        report += "\n█ نهاية التقرير".center(68) + "█"
        report += "\n█"*70 + "\n"
        
        return report
    
    def generate_dupont_report(self, dupont, waterfall=None, industry=None, recommendations=None):
        """توليد تقرير تحليل DuPont (نصي — يُصدَّر لاحقاً إلى PDF)"""

        _REC_TEXT = {
            'rec_npm_low': 'تحسين هامش الربح الصافي عبر خفض التكاليف أو رفع الأسعار (الهدف: ≥ 5%)',
            'rec_npm_ok': 'هامش الربح الصافي مقبول — العمل على تحسينه نحو 5-10%',
            'rec_npm_strong': 'هامش الربح الصافي ممتاز — حافظ على هذا الأداء',
            'rec_at_low': 'زيادة كفاءة استخدام الأصول لرفع معدل الدوران (الهدف: ≥ 1.0)',
            'rec_at_ok': 'معدل دوران الأصول مقبول — مراجعة السياسات التجارية للتحسين',
            'rec_at_strong': 'معدل دوران الأصول ممتاز — استغلال مثالي للأصول',
            'rec_em_high': 'الرافعة المالية مرتفعة — خفّض الديون لتقليل المخاطر (الهدف: ≤ 3)',
            'rec_em_ok': 'رافعة مالية متوازنة — هيكل تمويل سليم',
            'rec_em_low': 'رافعة مالية منخفضة — مجال لتحسين العائد عبر تمويل مدروس',
            'rec_roe_low': 'العائد على حقوق الملكية منخفض — راجع مكونات DuPont الثلاثة (الهدف: ≥ 10%)',
            'rec_roe_ok': 'العائد على حقوق الملكية مقبول — نحو التحسين المستمر',
            'rec_roe_strong': 'العائد على حقوق الملكية ممتاز — أداء متميز',
            'rec_industry_gap': 'فجوة مقارنة بمتوسط القطاع — ارفع القيمة نحو متوسط القطاع',
        }

        report = "\n" + "═"*70
        report += "\n📊 تقرير تحليل DuPont"
        report += f"\nالشركة: {self.company_name}"
        report += f"\nالسنة المالية: {self.fiscal_year}"
        report += f"\nتاريخ التقرير: {self.generated_date}"
        report += "\n" + "═"*70

        report += "\n\n🔢 مكونات DuPont:"
        report += "\n" + "-"*70
        report += f"\nهامش الربح الصافي (Net Profit Margin): {dupont.get('net_profit_margin', 0)}%"
        report += f"\nمعدل دوران الأصول (Asset Turnover): {dupont.get('asset_turnover', 0)}"
        report += f"\nالرافعة المالية (Equity Multiplier): {dupont.get('equity_multiplier', 0)}"
        report += f"\nالعائد على حقوق الملكية (ROE): {dupont.get('roe', 0)}%"

        if waterfall:
            report += "\n\n📐 تحليل الشلال (Waterfall):"
            report += "\n" + "-"*70
            report += f"\nمساهمة هامش الربح الأساسية: {waterfall.get('base', 0)}"
            report += f"\nأثر دوران الأصول: {waterfall.get('turnover_effect', 0)}"
            report += f"\nأثر الرافعة المالية: {waterfall.get('leverage_effect', 0)}"
            report += f"\nالإجمالي (ROE): {waterfall.get('total', 0)}"

        if industry:
            report += "\n\n🏭 مقارنة مع القطاع:"
            report += "\n" + "-"*70
            report += f"\n{'المكوّن':<25} {'الشركة':<12} {'متوسط القطاع':<14} {'الحالة'}"
            for component, cmp_data in industry.items():
                report += f"\n{component:<25} {cmp_data['company_value']:<12} {cmp_data['sector_average']:<14} {cmp_data['status']}"

        if recommendations:
            report += "\n\n🎯 التوصيات:"
            report += "\n" + "-"*70
            for rec in recommendations:
                text = _REC_TEXT.get(rec.get('code'), rec.get('code', ''))
                if rec.get('target') is not None and rec.get('code') == 'rec_industry_gap':
                    text += f" (المتوسط: {rec['target']})"
                elif rec.get('target') is not None:
                    text += f" — حالياً: {rec.get('company_value')}"
                report += f"\n• {text}"

        report += "\n\n" + "═"*70 + "\n"
        return report

    def generate_scenario_report(self, scenarios, comparison=None, sensitivity=None):
        """توليد تقرير تحليل السيناريوهات (نصي — يُصدَّر لاحقاً إلى PDF)"""

        _SCEN_LABELS = {
            'best': '🏆 الحالة المثالية (Best Case)',
            'base': '📊 الحالة الطبيعية (Base Case)',
            'worst': '⚠️ أسوأ حالة (Worst Case)',
        }
        _OUTCOME = {
            'profit': 'أرباح عالية',
            'loss': 'خسائر محتملة',
            'base': 'الأداء الحالي',
            'decline': 'انخفاض في الأرباح',
        }

        report = "\n" + "═"*70
        report += "\n🎭 تقرير تحليل السيناريوهات"
        report += f"\nالشركة: {self.company_name}"
        report += f"\nالسنة المالية: {self.fiscal_year}"
        report += f"\nتاريخ التقرير: {self.generated_date}"
        report += "\n" + "═"*70

        for sc_type in ("best", "base", "worst"):
            sc = scenarios.get(sc_type)
            if not sc:
                continue
            report += f"\n\n{_SCEN_LABELS.get(sc_type, sc_type)}"
            report += "\n" + "-"*70
            assumptions = sc.get("assumptions", {})
            report += f"\nتغيير المبيعات: {assumptions.get('revenue_change_pct', 0)*100:+.1f}%"
            report += f"\nتغيير التكاليف: {assumptions.get('cost_change_pct', 0)*100:+.1f}%"
            report += f"\nتغيير الكفاءة: {assumptions.get('efficiency_change_pct', 0)*100:+.1f}%"
            report += f"\nالإيرادات: {sc.get('revenue', 0):,.2f}"
            report += f"\nتكلفة البضاعة المباعة: {sc.get('cogs', 0):,.2f}"
            report += f"\nالمصاريف التشغيلية: {sc.get('operating_expenses', 0):,.2f}"
            report += f"\nصافي الربح: {sc.get('net_income', 0):,.2f}"
            report += f"\nهامش صافي الربح: {sc.get('net_profit_margin', 0):.2f}%"
            report += f"\nمعدل دوران الأصول: {sc.get('asset_turnover', 0):.4f}"
            report += f"\nالعائد على الأصول (ROA): {sc.get('roa', 0):.2f}%"
            report += f"\nالعائد على حقوق الملكية (ROE): {sc.get('roe', 0):.2f}%"
            report += f"\nالنتيجة: {_OUTCOME.get(sc.get('outcome'), sc.get('outcome', ''))}"

        if comparison:
            report += "\n\n📊 جدول المقارنة بين السيناريوهات"
            report += "\n" + "-"*70
            report += f"\n{'المؤشر':<25} {'مثالي':<14} {'طبيعي':<14} {'أسوأ':<14}"
            _METRIC_LABELS = {
                'revenue': 'الإيرادات', 'cogs': 'تكلفة البضاعة',
                'operating_expenses': 'المصاريف التشغيلية', 'gross_profit': 'الربح الإجمالي',
                'net_income': 'صافي الربح', 'net_profit_margin': 'هامش صافي الربح',
                'asset_turnover': 'دوران الأصول', 'roa': 'ROA', 'roe': 'ROE',
            }
            for metric, row in comparison.items():
                label = _METRIC_LABELS.get(metric, metric)
                report += f"\n{label:<25} {row['best']:<14} {row['base']:<14} {row['worst']:<14}"

        if sensitivity:
            report += "\n\n📈 تحليل الحساسية (تأثير تغيير المتغير)"
            report += "\n" + "-"*70
            report += f"\n{'نسبة التغيير':<15} {'صافي الربح':<18} {'هامش الربح':<15} {'ROE'}"
            for row in sensitivity:
                pct = row['pct_change'] * 100
                report += f"\n{pct:+.2f}%{'':<11} {row['net_income']:<18,.2f} {row['net_profit_margin']:<15.2f} {row['roe']:.2f}"

        report += "\n\n" + "═"*70 + "\n"
        return report

    def export_report_to_file(self, report_content, filename):
        """تصدير التقرير إلى ملف"""
        try:
            with open(filename, 'w', encoding='utf-8') as f:
                f.write(report_content)
            log.info("Report exported to: %s", filename)
            return True
        except Exception as e:
            log.error("Export error: %s", e)
            return False
    
    @staticmethod
    def _get_arabic_font_path():
        """الحصول على مسار الخط العربي"""
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        font_path = os.path.join(base, "ui", "resources", "fonts", "Amiri-Regular.ttf")
        if os.path.exists(font_path):
            return font_path
        return None

    @staticmethod
    def _get_arabic_bold_font_path():
        """الحصول على مسار الخط العربي العريض"""
        base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
        font_path = os.path.join(base, "ui", "resources", "fonts", "Amiri-Bold.ttf")
        if os.path.exists(font_path):
            return font_path
        return None

    def export_to_pdf(self, report_content, filename):
        try:
            from fpdf import FPDF

            class ArabicPDF(FPDF):
                def __init__(self, font_path=None, bold_path=None):
                    super().__init__()
                    self._font_path = font_path
                    self._bold_path = bold_path

                def header(self):
                    if self._font_path:
                        self.set_font("Amiri", size=9)
                    else:
                        self.set_font("Helvetica", size=9)
                    self.set_text_color(150, 150, 150)
                    self.cell(0, 8, "Smart Accounting Platform - Financial Report", ln=True, align="C")
                    self.set_draw_color(52, 152, 219)
                    self.set_line_width(0.5)
                    self.line(10, self.get_y(), 200, self.get_y())
                    self.ln(4)

                def footer(self):
                    self.set_y(-15)
                    if self._font_path:
                        self.set_font("Amiri", size=8)
                    else:
                        self.set_font("Helvetica", size=8)
                    self.set_text_color(150, 150, 150)
                    self.cell(0, 10, f"Page {self.page_no()}/{{nb}}  |  Generated by Smart Accounting Platform", align="C")

            font_path = self._get_arabic_font_path()
            bold_path = self._get_arabic_bold_font_path()

            has_arabic = any(
                '\u0600' <= c <= '\u06FF' or '\u0750' <= c <= '\u077F'
                for c in report_content
            )
            if has_arabic and not font_path:
                log.error(
                    "Cannot export PDF: Arabic content found but the Amiri font is "
                    "missing. Please restore ui/resources/fonts/Amiri-Regular.ttf."
                )
                return False

            pdf = ArabicPDF(font_path, bold_path)
            pdf.alias_nb_pages()
            pdf.set_auto_page_break(auto=True, margin=20)

            if font_path:
                pdf.add_font("Amiri", "", font_path, uni=True)
                if bold_path:
                    pdf.add_font("Amiri", "B", bold_path, uni=True)

            pdf.add_page()

            if self.company_name:
                if font_path:
                    pdf.set_font("Amiri", "B" if bold_path else "", 14)
                else:
                    pdf.set_font("Helvetica", "B", 14)
                pdf.set_text_color(41, 128, 185)
                pdf.cell(0, 10, self.company_name, ln=True, align="C")
                pdf.set_font("Amiri" if font_path else "Helvetica", "", 10)
                pdf.set_text_color(100, 100, 100)
                pdf.cell(0, 6, f"Fiscal Year: {self.fiscal_year}  |  Generated: {self.generated_date}", ln=True, align="C")
                pdf.set_draw_color(52, 152, 219)
                pdf.set_line_width(0.8)
                pdf.line(10, pdf.get_y() + 2, 200, pdf.get_y() + 2)
                pdf.ln(8)

            section_colors = {
                "═": (41, 128, 185),
                "█": (39, 174, 96),
                "▓": (231, 76, 60),
            }

            for line in report_content.split('\n'):
                stripped = line.strip()
                if not stripped:
                    pdf.ln(3)
                    continue

                is_section = any(c in stripped for c in ["═", "█", "▓"])
                is_separator = stripped.startswith("---") or stripped.startswith("===")

                if is_section:
                    r, g, b = (41, 128, 185)
                    for c, rgb in section_colors.items():
                        if c in stripped:
                            r, g, b = rgb
                            break
                    pdf.set_fill_color(r, g, b)
                    pdf.set_text_color(255, 255, 255)
                    font_style = "B" if bold_path else ""
                    if font_path:
                        pdf.set_font("Amiri", font_style, 11)
                    else:
                        pdf.set_font("Helvetica", font_style, 11)
                    pdf.cell(0, 9, stripped, ln=True, fill=True)
                    pdf.set_text_color(30, 30, 30)
                elif is_separator:
                    pdf.set_draw_color(200, 200, 200)
                    pdf.set_line_width(0.3)
                    pdf.line(10, pdf.get_y(), 200, pdf.get_y())
                    pdf.ln(3)
                else:
                    has_arabic = any('\u0600' <= c <= '\u06FF' or '\u0750' <= c <= '\u077F' for c in stripped)
                    is_bold = stripped.startswith("📊") or stripped.startswith("🎯") or stripped.startswith("📈")
                    if has_arabic and font_path:
                        style = "B" if is_bold and bold_path else ""
                        pdf.set_font("Amiri", style, 10)
                    else:
                        style = "B" if is_bold else ""
                        pdf.set_font("Helvetica", style, 10)
                    pdf.set_text_color(30, 30, 30)
                    pdf.cell(0, 7, stripped, ln=True)

            pdf.output(filename)
            return True
        except Exception as e:
            log.error("PDF export error: %s", e)
            return False

    def export_to_excel(self, filename, financial_data=None, ratios=None, tax_data=None):
        """تصدير التقرير إلى Excel مع رسوم بيانية"""
        try:
            from openpyxl import Workbook
            from openpyxl.chart import BarChart, PieChart, Reference
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()

            header_font = Font(bold=True, size=12, color="FFFFFF")
            header_fill = PatternFill(start_color="2980B9", end_color="2980B9", fill_type="solid")
            title_font = Font(bold=True, size=14)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            ws = wb.active
            ws.title = "Financial Report"

            ws.merge_cells('A1:D1')
            ws['A1'] = f"Financial Report - {self.company_name}"
            ws['A1'].font = title_font
            ws['A2'] = f"Fiscal Year: {self.fiscal_year}"
            ws['A3'] = f"Generated: {self.generated_date}"

            row = 5
            if financial_data:
                ws.cell(row=row, column=1, value="Balance Sheet").font = Font(bold=True, size=12)
                row += 1
                headers = ["Item", "Value"]
                for col, h in enumerate(headers, 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = header_font
                    c.fill = header_fill
                    c.border = thin_border
                row += 1

                bs_items = [
                    ("Current Assets", financial_data.get('current_assets', 0)),
                    ("Inventory", financial_data.get('inventory', 0)),
                    ("Total Assets", financial_data.get('total_assets', 0)),
                    ("Current Liabilities", financial_data.get('current_liabilities', 0)),
                    ("Total Liabilities", financial_data.get('total_liabilities', 0)),
                    ("Equity", financial_data.get('equity', 0)),
                ]
                chart_start = row
                for label, val in bs_items:
                    ws.cell(row=row, column=1, value=label).border = thin_border
                    c = ws.cell(row=row, column=2, value=val)
                    c.number_format = '#,##0.00'
                    c.border = thin_border
                    row += 1
                chart_end = row - 1

                chart = BarChart()
                chart.title = "Balance Sheet Structure"
                chart.y_axis.title = "Amount (DZD)"
                data_ref = Reference(ws, min_col=2, min_row=chart_start - 1, max_row=chart_end)
                cats_ref = Reference(ws, min_col=1, min_row=chart_start, max_row=chart_end)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cats_ref)
                chart.width = 20
                chart.height = 12
                ws.add_chart(chart, f"D5")

                row += 1

            if financial_data:
                ws.cell(row=row, column=1, value="Income Statement").font = Font(bold=True, size=12)
                row += 1
                for col, h in enumerate(["Item", "Value"], 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = header_font
                    c.fill = header_fill
                    c.border = thin_border
                row += 1

                income_items = [
                    ("Revenue", financial_data.get('revenue', 0)),
                    ("COGS", financial_data.get('cost_of_goods_sold', 0)),
                    ("Gross Profit", financial_data.get('gross_profit', 0)),
                    ("Net Income", financial_data.get('net_income', 0)),
                ]
                for label, val in income_items:
                    ws.cell(row=row, column=1, value=label).border = thin_border
                    c = ws.cell(row=row, column=2, value=val)
                    c.number_format = '#,##0.00'
                    c.border = thin_border
                    row += 1
                row += 1

            if ratios:
                ws.cell(row=row, column=1, value="Financial Ratios").font = Font(bold=True, size=12)
                row += 1
                for col, h in enumerate(["Ratio", "Value"], 1):
                    c = ws.cell(row=row, column=col, value=h)
                    c.font = header_font
                    c.fill = header_fill
                    c.border = thin_border
                row += 1

                ratio_labels = {
                    'current_ratio': 'Current Ratio',
                    'quick_ratio': 'Quick Ratio',
                    'gross_profit_margin': 'Gross Profit Margin (%)',
                    'net_profit_margin': 'Net Profit Margin (%)',
                    'roa': 'ROA (%)',
                    'roe': 'ROE (%)',
                    'asset_turnover': 'Asset Turnover',
                    'debt_to_equity': 'Debt/Equity',
                    'debt_ratio': 'Debt Ratio',
                }
                chart_start = row
                for key, label in ratio_labels.items():
                    val = ratios.get(key, 0)
                    ws.cell(row=row, column=1, value=label).border = thin_border
                    c = ws.cell(row=row, column=2, value=val)
                    c.number_format = '0.00'
                    c.border = thin_border
                    row += 1
                chart_end = row - 1

                chart2 = BarChart()
                chart2.title = "Key Financial Ratios"
                data_ref2 = Reference(ws, min_col=2, min_row=chart_start - 1, max_row=chart_end)
                cats_ref2 = Reference(ws, min_col=1, min_row=chart_start, max_row=chart_end)
                chart2.add_data(data_ref2, titles_from_data=True)
                chart2.set_categories(cats_ref2)
                chart2.width = 20
                chart2.height = 12
                ws.add_chart(chart2, f"D{chart_start}")

            for col_letter in ['A', 'B']:
                ws.column_dimensions[col_letter].width = 30

            wb.save(filename)
            log.info("Excel exported to: %s", filename)
            return True
        except Exception as e:
            log.error("Excel export error: %s", e)
            return False