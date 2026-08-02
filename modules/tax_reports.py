# قوالب الإقرارات الجبائية الجزائرية — DGI
# =========================================

import os
from datetime import datetime

from utils.app_logger import get_logger

logger = get_logger("tax_reports")


class TaxDeclarationGenerator:
    """مولّد قوالب الإقرارات الجبائية الجاهزة للإيداع لدى DGI

    يدعم:
    - G N°50: إقرار TVA الشهري
    - G N°57: الإقرار السنوي IBS
    - DAS: الإقرار السنوي للأجور
    """

    DECLARATION_TYPES = {
        "g50": {
            "name_ar": "G N°50 — إقرار TVA الشهري",
            "name_en": "G N°50 — Monthly VAT Declaration",
            "name_fr": "G N°50 — Déclaration TVA mensuelle",
            "deadline_ar": "قبل يوم 20 من الشهر الموالي",
        },
        "g57": {
            "name_ar": "G N°57 — الإقرار السنوي IBS",
            "name_en": "G N°57 — Annual IBS Declaration",
            "name_fr": "G N°57 — Déclaration annuelle IBS",
            "deadline_ar": "قبل 30 أبريل من السنة الموالية",
        },
        "das": {
            "name_ar": "DAS — الإقرار السنوي للأجور",
            "name_en": "DAS — Annual Wages Declaration",
            "name_fr": "DAS — Déclaration annuelle des salaires",
            "deadline_ar": "قبل 31 يناير من السنة الموالية",
        },
    }

    def __init__(self, engine=None):
        from modules.tax import TaxEngine
        self.engine = engine or TaxEngine()

    def get_declaration_types(self):
        """قائمة أنواع الإقرارات المتوفرة"""
        return list(self.DECLARATION_TYPES.keys())

    def get_declaration_info(self, decl_type):
        """معلومات نوع الإقرار"""
        return self.DECLARATION_TYPES.get(decl_type, {})

    # ==================== البناء ====================

    def build_header(self, company_info, fiscal_year, period_label=""):
        """بناء ترويسة الإقرار من بيانات الشركة"""
        return {
            "company_name": company_info.get("company_name", ""),
            "nif": company_info.get("nif", ""),
            "rc": company_info.get("rc", ""),
            "ai": company_info.get("ai", ""),
            "address": company_info.get("address", ""),
            "dgi_center": company_info.get("dgi_center", ""),
            "fiscal_year": fiscal_year,
            "period": period_label,
        }

    def generate_g50(self, header, month, year, monthly_turnover,
                     tva_collected, tva_deductible, previous_credit=0):
        """بناء إقرار G N°50 (TVA الشهري)"""
        net_tva = self.engine.calculate_tva_refund(
            tva_collected, tva_deductible, previous_credit
        )
        return {
            "type": "g50",
            "header": header,
            "period": {"month": month, "year": year},
            "turnover": round(monthly_turnover, 2),
            "rates": self.engine.get_tva_rates(),
            "tva_collected": round(tva_collected, 2),
            "tva_deductible": round(tva_deductible, 2),
            "net_tva": net_tva,
        }

    def generate_g57(self, header, taxable_income, acomptes_paid=0,
                     activity_type="other", reinvestment=0):
        """بناء إقرار G N°57 (الإقرار السنوي IBS)"""
        ibs = self.engine.calculate_ibs(taxable_income, activity_type)
        acomptes = self.engine.calculate_ibs_acomptes(taxable_income, activity_type)
        balance = self.engine.calculate_ibs_balance(
            taxable_income, activity_type, acomptes_paid
        )
        return {
            "type": "g57",
            "header": header,
            "taxable_income": round(taxable_income, 2),
            "ibs": ibs,
            "acomptes": acomptes,
            "balance": balance,
            "reinvestment": round(reinvestment, 2),
        }

    def generate_das(self, header, monthly_payroll, number_of_employees, avg_salary=0):
        """بناء إقرار DAS (الإقرار السنوي للأجور)"""
        das = self.engine.build_das_data(
            monthly_payroll, number_of_employees, avg_salary
        )
        return {
            "type": "das",
            "header": header,
            "data": das,
        }

    def generate(self, decl_type, data):
        """بناء الإقرار حسب نوعه"""
        header = data.get("header", {})
        if decl_type == "g50":
            return self.generate_g50(
                header,
                data.get("month", 1),
                data.get("year", datetime.now().year),
                data.get("monthly_turnover", 0),
                data.get("tva_collected", 0),
                data.get("tva_deductible", 0),
                data.get("previous_credit", 0),
            )
        if decl_type == "g57":
            return self.generate_g57(
                header,
                data.get("taxable_income", 0),
                data.get("acomptes_paid", 0),
                data.get("activity_type", "other"),
                data.get("reinvestment", 0),
            )
        if decl_type == "das":
            return self.generate_das(
                header,
                data.get("monthly_payroll", 0),
                data.get("number_of_employees", 0),
                data.get("avg_salary", 0),
            )
        raise ValueError(f"Unknown declaration type: {decl_type}")

    # ==================== العرض ====================

    def _fmt(self, amount):
        return f"{amount:,.2f} DZD"

    def _header_block(self, declaration):
        header = declaration.get("header", {})
        info = self.DECLARATION_TYPES.get(declaration["type"], {})
        lines = [
            "═" * 70,
            f"█ {info.get('name_ar', 'إقرار جبائي')}",
            "═" * 70,
            f"الشركة: {header.get('company_name', '')}",
            f"NIF: {header.get('nif', '')} | RC: {header.get('rc', '')} | AI: {header.get('ai', '')}",
            f"العنوان: {header.get('address', '')}",
            f"المديرية الجبائية: {header.get('dgi_center', '')}",
            "-" * 70,
        ]
        period = header.get("period", "")
        if period:
            lines.append(f"الفترة: {period}")
        else:
            lines.append(f"السنة المالية: {header.get('fiscal_year', '')}")
        lines.append("-" * 70)
        return lines

    def _footer_block(self, declaration):
        info = self.DECLARATION_TYPES.get(declaration["type"], {})
        return [
            "-" * 70,
            "توقيع المدير والممثل القانوني: ............................",
            "",
            f"أجل الإيداع: {info.get('deadline_ar', '')}",
            "يودع الإقرار لدى المديرية الجبائية المختصة إقليمياً",
            "═" * 70,
        ]

    def _render_g50(self, declaration):
        lines = self._header_block(declaration)
        period = declaration["period"]
        lines += [
            f"رقم الأعمال الشهري (بدون TVA): {self._fmt(declaration['turnover'])}",
            f"الشهر: {period['month']} | السنة: {period['year']}",
            "-" * 70,
            f"TVA المحصلة: {self._fmt(declaration['tva_collected'])}",
            f"TVA القابلة للخصم: {self._fmt(declaration['tva_deductible'])}",
        ]
        net = declaration["net_tva"]
        if net["status"] == "payable":
            lines.append(f"المبلغ الواجب دفعه: {self._fmt(net['net_payable'])}")
        else:
            lines.append(f"رصيد يُرحل/يُسترجع: {self._fmt(net['remaining_credit'])}")
        lines.append(f"رصيد سابق: {self._fmt(net['previous_credit'])}")
        lines += self._footer_block(declaration)
        return "\n".join(lines)

    def _render_g57(self, declaration):
        lines = self._header_block(declaration)
        ibs = declaration["ibs"]
        acomptes = declaration["acomptes"]
        balance = declaration["balance"]
        lines += [
            f"النتيجة الخاضعة للضريبة: {self._fmt(declaration['taxable_income'])}",
            f"نسبة IBS: {ibs['rate_used'] * 100:.0f}%",
            f"ضريبة IBS: {self._fmt(ibs['tax_amount'])}",
            f"الحد الأدنى مطبق: {'نعم' if ibs['minimum_applied'] else 'لا'}",
            "-" * 70,
            "الدفعات المقدمة (الأقساط):",
        ]
        for a in acomptes["acomptes"]:
            lines.append(f"  {a['label']} (شهر {a['month']}): {self._fmt(a['amount'])}")
        lines += [
            f"إجمالي الأقساط: {self._fmt(acomptes['total_acomptes'])}",
            f"الأقساط المدفوعة: {self._fmt(balance['acomptes_paid'])}",
            "-" * 70,
            f"الباقي المستحق: {self._fmt(balance['balance_due'])}",
        ]
        if balance["refund_amount"] > 0:
            lines.append(f"فائض يُسترجع: {self._fmt(balance['refund_amount'])}")
        lines += self._footer_block(declaration)
        return "\n".join(lines)

    def _render_das(self, declaration):
        lines = self._header_block(declaration)
        d = declaration["data"]
        lines += [
            f"عدد الموظفين: {d['number_of_employees']}",
            f"كتلة الأجور السنوية: {self._fmt(d['annual_payroll'])}",
            "-" * 70,
            f"CNAS — حصة صاحب العمل (سنوي): {self._fmt(d['cnas_employer_annual'])}",
            f"CNAS — حصة الموظف (سنوي): {self._fmt(d['cnas_employee_annual'])}",
            f"CNAC — حصة صاحب العمل (سنوي): {self._fmt(d['cnac_employer_annual'])}",
            f"CNAC — حصة الموظف (سنوي): {self._fmt(d['cnac_employee_annual'])}",
            f"IRG المقتطع من الأجور (سنوي): {self._fmt(d['irg_withheld_annual'])}",
            "-" * 70,
            f"صافي الأجور المدفوعة: {self._fmt(d['net_payroll_annual'])}",
        ]
        lines += self._footer_block(declaration)
        return "\n".join(lines)

    def render_text(self, declaration):
        """عرض الإقرار كنص منسّق (متوافق مع تصدير PDF)"""
        renderers = {
            "g50": self._render_g50,
            "g57": self._render_g57,
            "das": self._render_das,
        }
        renderer = renderers.get(declaration.get("type"))
        if not renderer:
            raise ValueError(f"Unknown declaration type: {declaration.get('type')}")
        return renderer(declaration)

    # ==================== التصدير ====================

    def export_pdf(self, declaration, filename):
        """تصدير الإقرار إلى PDF (عبر ReportGenerator)"""
        try:
            from modules.reporting import ReportGenerator
            text = self.render_text(declaration)
            header = declaration.get("header", {})
            generator = ReportGenerator(
                header.get("company_name", ""),
                header.get("fiscal_year", ""),
            )
            return generator.export_to_pdf(text, filename)
        except Exception as e:
            logger.error(f"Declaration PDF export failed: {e}")
            return False

    def export_excel(self, declaration, filename):
        """تصدير الإقرار إلى Excel بصيغة رسمية"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            info = self.DECLARATION_TYPES.get(declaration["type"], {})
            header = declaration.get("header", {})
            wb = Workbook()
            ws = wb.active
            ws.title = "Declaration"

            title_font = Font(bold=True, size=14, color="1F4E79")
            sub_font = Font(bold=True, size=11)
            label_font = Font(bold=True)
            thin = Side(style="thin", color="BBBBBB")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            fill = PatternFill("solid", fgColor="D9E2F3")

            ws.merge_cells("A1:D1")
            ws["A1"] = header.get("company_name", "")
            ws["A1"].font = title_font
            ws["A1"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A2:D2")
            ws["A2"] = info.get("name_ar", "")
            ws["A2"].font = sub_font
            ws["A2"].alignment = Alignment(horizontal="center")

            ws.merge_cells("A3:D3")
            period = header.get("period", "")
            ws["A3"] = period or f"السنة المالية: {header.get('fiscal_year', '')}"
            ws["A3"].alignment = Alignment(horizontal="center")

            row = 5
            identity = [
                ("NIF", header.get("nif", "")),
                ("RC", header.get("rc", "")),
                ("AI", header.get("ai", "")),
                ("العنوان", header.get("address", "")),
                ("المديرية الجبائية", header.get("dgi_center", "")),
            ]
            for label, value in identity:
                ws.cell(row=row, column=1, value=label).font = label_font
                ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
                ws.cell(row=row, column=2, value=value)
                row += 1

            row += 1
            ws.cell(row=row, column=1, value="البيان").font = sub_font
            ws.cell(row=row, column=2, value="القيمة (DZD)").font = sub_font
            for col in range(1, 5):
                ws.cell(row=row, column=col).fill = fill
                ws.cell(row=row, column=col).border = border
            row += 1

            def add_row(label, value):
                nonlocal row
                ws.cell(row=row, column=1, value=label).font = label_font
                ws.cell(row=row, column=2, value=value)
                ws.cell(row=row, column=2).number_format = "#,##0.00"
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = border
                row += 1

            if declaration["type"] == "g50":
                period = declaration["period"]
                add_row("الشهر", period["month"])
                add_row("السنة", period["year"])
                add_row("رقم الأعمال الشهري (بدون TVA)", declaration["turnover"])
                add_row("TVA المحصلة", declaration["tva_collected"])
                add_row("TVA القابلة للخصم", declaration["tva_deductible"])
                add_row("رصيد سابق", declaration["net_tva"]["previous_credit"])
                if declaration["net_tva"]["status"] == "payable":
                    add_row("المبلغ الواجب دفعه", declaration["net_tva"]["net_payable"])
                else:
                    add_row("رصيد يُرحل/يُسترجع", declaration["net_tva"]["remaining_credit"])
            elif declaration["type"] == "g57":
                ibs = declaration["ibs"]
                balance = declaration["balance"]
                add_row("النتيجة الخاضعة للضريبة", declaration["taxable_income"])
                add_row("نسبة IBS", f"{ibs['rate_used'] * 100:.0f}%")
                add_row("ضريبة IBS", ibs["tax_amount"])
                add_row("إجمالي الأقساط", declaration["acomptes"]["total_acomptes"])
                add_row("الأقساط المدفوعة", balance["acomptes_paid"])
                add_row("الباقي المستحق", balance["balance_due"])
                if balance["refund_amount"] > 0:
                    add_row("فائض يُسترجع", balance["refund_amount"])
            else:
                d = declaration["data"]
                add_row("عدد الموظفين", d["number_of_employees"])
                add_row("كتلة الأجور السنوية", d["annual_payroll"])
                add_row("CNAS — حصة صاحب العمل (سنوي)", d["cnas_employer_annual"])
                add_row("CNAS — حصة الموظف (سنوي)", d["cnas_employee_annual"])
                add_row("CNAC — حصة صاحب العمل (سنوي)", d["cnac_employer_annual"])
                add_row("CNAC — حصة الموظف (سنوي)", d["cnac_employee_annual"])
                add_row("IRG المقتطع من الأجور (سنوي)", d["irg_withheld_annual"])
                add_row("صافي الأجور المدفوعة", d["net_payroll_annual"])

            row += 1
            ws.cell(row=row, column=1, value="توقيع المدير والممثل القانوني: ....................").font = sub_font
            row += 1
            ws.cell(row=row, column=1, value=f"أجل الإيداع: {info.get('deadline_ar', '')}")
            row += 1
            ws.cell(row=row, column=1, value="يودع الإقرار لدى المديرية الجبائية المختصة إقليمياً")

            for col in range(1, 5):
                ws.column_dimensions[chr(64 + col)].width = 26

            wb.save(filename)
            return True
        except Exception as e:
            logger.error(f"Declaration Excel export failed: {e}")
            return False


tax_declaration_generator = TaxDeclarationGenerator()
